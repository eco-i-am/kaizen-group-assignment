import pandas as pd
import openpyxl
from collections import defaultdict

def analyze_output():
    """Analyze the output file to see what users are included"""
    
    input_file = 'sample_merged_data.xlsx'
    output_file = 'grouped_participants.xlsx'
    
    # Read input data
    try:
        df_input = pd.read_excel(input_file, sheet_name='Merged Data')
        print(f"✅ Input file: {len(df_input)} records")
    except Exception as e:
        print(f"❌ Error reading input file: {e}")
        return
    
    # Read output file
    try:
        wb_output = openpyxl.load_workbook(output_file)
        ws_output = wb_output.active
        print(f"✅ Output file: {ws_output.max_row - 1} rows (excluding header)")
    except Exception as e:
        print(f"❌ Error reading output file: {e}")
        return
    
    # Extract user IDs from input
    input_user_ids = set()
    for _, row in df_input.iterrows():
        user_id = row.get('user_id', '')
        if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
            input_user_ids.add(str(user_id).strip())
    
    print(f"📊 Input user IDs: {len(input_user_ids)}")
    print(f"  Sample: {list(input_user_ids)[:10]}")
    
    # Extract user IDs from output
    output_user_ids = set()
    group_types = defaultdict(int)
    
    for row in range(2, ws_output.max_row + 1):  # Skip header
        group_name = ws_output.cell(row=row, column=1).value
        if not group_name:
            continue
            
        # Count group types
        if 'Solo' in str(group_name):
            group_types['solo'] += 1
        elif 'Requested Group' in str(group_name):
            group_types['requested'] += 1
        elif 'Team Group' in str(group_name):
            group_types['team'] += 1
        elif 'Excluded' in str(group_name):
            group_types['excluded'] += 1
        else:
            group_types['regular'] += 1
        
        # Extract user IDs from columns 2, 5, 8, 11, 14, 17, 20 (User ID columns)
        for col in [2, 5, 8, 11, 14, 17, 20]:
            user_id = ws_output.cell(row=row, column=col).value
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                output_user_ids.add(str(user_id).strip())
    
    print(f"📊 Output user IDs: {len(output_user_ids)}")
    print(f"  Sample: {list(output_user_ids)[:10]}")
    
    # Compare
    missing_in_output = input_user_ids - output_user_ids
    extra_in_output = output_user_ids - input_user_ids
    
    print(f"\n📊 COMPARISON:")
    print(f"  Users in input but missing from output: {len(missing_in_output)}")
    if missing_in_output:
        print(f"    Missing: {list(missing_in_output)}")
    
    print(f"  Users in output but not in input: {len(extra_in_output)}")
    if extra_in_output:
        print(f"    Extra: {list(extra_in_output)}")
    
    print(f"  Users in both: {len(input_user_ids & output_user_ids)}")
    
    # Group breakdown
    print(f"\n📋 GROUP BREAKDOWN:")
    for group_type, count in group_types.items():
        print(f"  {group_type}: {count} groups")
    
    # Check specific groups
    print(f"\n🔍 DETAILED GROUP ANALYSIS:")
    
    solo_users = set()
    requested_users = set()
    team_users = set()
    excluded_users = set()
    regular_users = set()
    
    for row in range(2, ws_output.max_row + 1):
        group_name = ws_output.cell(row=row, column=1).value
        if not group_name:
            continue
        
        # Extract user IDs for this group
        group_user_ids = []
        for col in [2, 5, 8, 11, 14, 17, 20]:
            user_id = ws_output.cell(row=row, column=col).value
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                group_user_ids.append(str(user_id).strip())
        
        # Categorize
        if 'Solo' in str(group_name):
            solo_users.update(group_user_ids)
        elif 'Requested Group' in str(group_name):
            requested_users.update(group_user_ids)
        elif 'Team Group' in str(group_name):
            team_users.update(group_user_ids)
        elif 'Excluded' in str(group_name):
            excluded_users.update(group_user_ids)
        else:
            regular_users.update(group_user_ids)
    
    print(f"  Solo users: {len(solo_users)}")
    print(f"  Requested group users: {len(requested_users)}")
    print(f"  Team group users: {len(team_users)}")
    print(f"  Excluded users: {len(excluded_users)}")
    print(f"  Regular group users: {len(regular_users)}")
    
    # Check for duplicates
    all_output_users = solo_users | requested_users | team_users | excluded_users | regular_users
    print(f"  Total unique users in output: {len(all_output_users)}")
    
    # Check for overlaps
    overlaps = []
    if solo_users & requested_users:
        overlaps.append(f"Solo & Requested: {len(solo_users & requested_users)}")
    if solo_users & team_users:
        overlaps.append(f"Solo & Team: {len(solo_users & team_users)}")
    if solo_users & regular_users:
        overlaps.append(f"Solo & Regular: {len(solo_users & regular_users)}")
    if requested_users & team_users:
        overlaps.append(f"Requested & Team: {len(requested_users & team_users)}")
    if requested_users & regular_users:
        overlaps.append(f"Requested & Regular: {len(requested_users & regular_users)}")
    if team_users & regular_users:
        overlaps.append(f"Team & Regular: {len(team_users & regular_users)}")
    
    if overlaps:
        print(f"  ⚠️  User overlaps detected:")
        for overlap in overlaps:
            print(f"    {overlap}")
    else:
        print(f"  ✅ No user overlaps detected")
    
    # Show some example groups
    print(f"\n📋 EXAMPLE GROUPS:")
    count = 0
    for row in range(2, ws_output.max_row + 1):
        if count >= 5:  # Show first 5 groups
            break
        group_name = ws_output.cell(row=row, column=1).value
        if group_name:
            group_user_ids = []
            for col in [2, 5, 8, 11, 14, 17, 20]:
                user_id = ws_output.cell(row=row, column=col).value
                if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                    group_user_ids.append(str(user_id).strip())
            
            print(f"  {group_name}: {len(group_user_ids)} users - {group_user_ids}")
            count += 1

if __name__ == "__main__":
    analyze_output() 