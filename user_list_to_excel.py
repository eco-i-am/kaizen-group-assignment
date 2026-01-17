import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import re

def safe_get_value(data_dict, key, default=''):
    """Safely get a value from dictionary, treating NaN as empty string"""
    if not data_dict or key not in data_dict:
        return default
    value = data_dict[key]
    if pd.isna(value) or value is None:
        return default
    return str(value).strip()

# File paths - Same as group_assignment_to_excel.py
INPUT_FILE = 'merged_users_grouping_preferences_20250719_133755.xlsx'  # Change this to your merged file
OUTPUT_FILE = 'user_list.xlsx'

# Column mapping for merged data (same as group_assignment_to_excel.py)
EXPECTED_COLUMNS = {
    'user_id': ['user_id', 'id', 'userid', 'user id', 'id_y', 'id_x'],
    'name': ['name', 'full_name', 'fullname', 'first_name', 'last_name', 'firstName', 'lastName'],
    'gender_identity': ['gender_identity', 'gender', 'genderidentity', 'genderIdentity'],
    'sex': ['sex', 'biological_sex', 'biologicalsex'],
    'residing_ph': ['residing_ph', 'residing_in_philippines', 'philippines_resident', 'lingInPhilippineExperience', 'residingInPhilippines'],
    'gender_preference': ['gender_preference', 'grouping_preference', 'preference', 'genderPref', 'groupGenderPreference'],
    'country': ['country', 'nationality'],
    'province': ['province', 'state_province'],
    'city': ['city', 'municipality'],
    'state': ['state', 'region', 'stat'],
    'go_solo': ['go_solo', 'solo', 'prefer_solo', 'goSolo'],
    'joining_as_student': ['joining_as_student', 'joiningAsStudent', 'student', 'is_student'],
    'kaizen_client_type': ['kaizen_client_type', 'kaizenClientType', 'client_type', 'clientType'],
    'accountability_buddies': ['accountability_buddies', 'accountabilityBuddies', 'accountability_buddies', 'buddies'],
    'has_accountability_buddies': ['has_accountability_buddies', 'hasAccountabilityBuddies', 'has_buddies'],
    'email': ['email', 'user_email', 'email_address', 'useremail', 'userEmail'],
    'temporary_team_name': ['temporaryTeamName', 'temporary_team_name', 'temp_team_name', 'team_name'],
    'previous_coach_name': ['previousCoachName', 'previous_coach_name', 'prev_coach_name', 'coach_name'],
    'current_goal': ['currentGoal', 'current_goal', 'goal'],
    'age_group': ['ageGroup', 'age_group', 'age']
}

# Helper for color coding based on sex
SEX_COLOR = {
    'male': 'ADD8E6',    # Blue
    'female': 'FFC0CB',  # Pink
}

# Maroon color for LGBTQ+ font
LGBTQ_FONT_COLOR = '800000'  # Maroon

# Light Green color for get_bigger goal IDs
GREEN_COLOR = '90EE90'  # Light Green

def format_location_display(member, column_mapping):
    """Format location display based on residing_ph status with enhanced logic"""
    residing_ph = str(member.get(column_mapping.get('residing_ph'), '0')).strip().lower()

    if residing_ph in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
        # Philippines resident - show "city, province" format
        city = member.get(column_mapping.get('city'), '')
        province = member.get(column_mapping.get('province'), '')

        # Use "MM" as acronym for Metro Manila
        if province and str(province).lower() == 'metro manila':
            province = 'MM'

        # Convert to strings to handle float values
        city_str = str(city) if city else ''
        province_str = str(province) if province else ''

        if city_str and province_str:
            return f"{city_str}, {province_str}"
        elif city_str:
            return city_str
        elif province_str:
            return province_str
        else:
            return ''
    else:
        # International resident - show enhanced format: international_city, international_state, location_identifier, country
        # Check if international-specific columns exist, otherwise fall back to regular columns
        international_city = member.get('internationalCity', '') or member.get(column_mapping.get('city'), '')
        international_state = member.get('internationalState', '') or member.get(column_mapping.get('state'), '')
        location_identifier = member.get('locationIdentifier', '') or member.get('location_identifier', '')
        country = member.get(column_mapping.get('country'), '')

        parts = []
        if international_city:
            parts.append(str(international_city))
        if international_state:
            parts.append(str(international_state))
        if location_identifier:
            parts.append(str(location_identifier))
        if country:
            parts.append(str(country))

        return ', '.join(parts) if parts else ''

def find_column_mapping(df):
    """Dynamically find column mapping based on available columns"""
    mapping = {}
    available_columns = [col.lower() for col in df.columns]

    for expected_key, possible_names in EXPECTED_COLUMNS.items():
        found = False
        for possible_name in possible_names:
            if possible_name.lower() in available_columns:
                # Find the exact column name (case-insensitive match)
                for col in df.columns:
                    if col.lower() == possible_name.lower():
                        mapping[expected_key] = col
                        found = True
                        break
                if found:
                    break

        if not found:
            mapping[expected_key] = None

    return mapping

def apply_color_to_cell(cell, sex, gender_identity=None, gender_preference=None, has_accountability_buddies=None, current_goal=None, is_user_id=False):
    """Apply color coding based on sex, font styling, and special fill coloring"""
    # Apply fill color based on sex (default)
    fill_color = None
    sex_lower = str(sex).lower().strip() if sex else ''
    if sex_lower in SEX_COLOR:
        fill_color = SEX_COLOR[sex_lower]

    # Special fill color for get_bigger goal (overrides sex color for User ID)
    if is_user_id and current_goal and str(current_goal).lower() == 'get_bigger':
        fill_color = GREEN_COLOR

    # Apply fill color if set
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Apply font formatting
    font_color = None
    is_bold = False
    is_underline = False

    # Maroon font color for LGBTQ+
    if gender_identity and str(gender_identity).lower() in ['lgbtq+', 'lgbtq']:
        font_color = LGBTQ_FONT_COLOR

    # Bold text for same_gender preference
    if gender_preference and str(gender_preference).lower().strip() == 'same_gender':
        is_bold = True

    # Underlined text for users with accountability buddies
    if has_accountability_buddies and str(has_accountability_buddies).lower() in ['1', '1.0', 'true', 'yes']:
        is_underline = True

    # Create font style - always create a new font with the desired properties
    cell.font = Font(
        color=font_color,
        bold=is_bold,
        underline='single' if is_underline else None
    )

def format_name_display(name, kaizen_client_type):
    """Format name with prefixes/suffixes based on kaizen_client_type"""
    if not name:
        return name

    formatted_name = str(name).strip()

    if kaizen_client_type:
        client_type = str(kaizen_client_type).lower().strip()

        # Add ** before name for team_member
        if client_type == "team_member":
            formatted_name = f"**{formatted_name}"

        # Add * at end for returning_s7
        elif client_type == "returning_latest":
            formatted_name = f"{formatted_name}*"

        # Add ** at end for returning_other
        elif client_type == "returning_other":
            formatted_name = f"{formatted_name}**"

    return formatted_name

def save_user_list_to_excel(data, filename_or_buffer, column_mapping):
    """Save user list to Excel with 6 columns: User ID, Name, Location, Coach and Age, Email Address, Group Mates"""
    wb = Workbook()
    ws = wb.active
    ws.title = "User List"

    # Header with 6 columns
    ws.append(["User ID", "Name", "Location", "Coach and Age", "Email Address", "Group Mates"])

    # Remove duplicates based on user_id before processing
    user_id_col = column_mapping.get('user_id')
    if user_id_col:
        unique_data = []
        seen_user_ids = set()
        duplicates_removed = 0

        for record in data:
            user_id_raw = record.get(user_id_col, '')
            user_id = str(user_id_raw).strip() if user_id_raw is not None else ''
            if user_id and user_id not in seen_user_ids:
                unique_data.append(record)
                seen_user_ids.add(user_id)
            elif user_id and user_id in seen_user_ids:
                duplicates_removed += 1
                # Skip duplicate
                continue
            elif not user_id:  # Keep records without user_id (though they should have one)
                unique_data.append(record)

        data = unique_data
        if duplicates_removed > 0:
            print(f"🧹 Deduplication: Removed {duplicates_removed} duplicate users, kept {len(data)} unique users")
        else:
            print(f"✅ No duplicates found: {len(data)} unique users")
    else:
        print("⚠️ No user_id column found - cannot deduplicate")

    # Sort data by user_id, name, city, coach for consistent ordering
    sorted_data = sorted(data, key=lambda m: (
        m.get(column_mapping.get('user_id'), ''),
        m.get(column_mapping.get('name'), ''),
        m.get(column_mapping.get('city'), ''),
        m.get(column_mapping.get('previous_coach_name'), '')
    ))

    # Add each user as a row
    for member in sorted_data:
        location_display = format_location_display(member, column_mapping)
        coach_name = safe_get_value(member, column_mapping.get('previous_coach_name', ''), '')
        age_group = safe_get_value(member, column_mapping.get('age_group', ''), '')
        kaizen_client_type = safe_get_value(member, column_mapping.get('kaizen_client_type', ''), '')
        sex = safe_get_value(member, column_mapping.get('sex', ''), '')
        gender_identity = safe_get_value(member, column_mapping.get('gender_identity', ''), '')
        gender_preference = safe_get_value(member, column_mapping.get('gender_preference', ''), '')
        has_accountability_buddies = safe_get_value(member, column_mapping.get('has_accountability_buddies', ''), '')
        current_goal = safe_get_value(member, column_mapping.get('current_goal', ''), '')

        # Format coach name with age group in parentheses
        # If coach_name is blank (NaN), keep it blank regardless of age_group
        if coach_name:
            coach_with_age = coach_name
            if age_group:
                coach_with_age = f"{coach_name} ({age_group})"
        else:
            if age_group:
                coach_with_age = f"({age_group})"
            else:
                coach_with_age = ""  # Keep blank if coach is NaN

        # Format the name with prefixes/suffixes based on kaizen_client_type
        original_name = member.get(column_mapping.get('name'), '')
        formatted_name = format_name_display(original_name, kaizen_client_type)

        row = [
            safe_get_value(member, column_mapping.get('user_id', ''), ''),
            formatted_name,  # Use formatted name with prefixes/suffixes
            location_display,
            coach_with_age,  # Use formatted coach name with age group
            safe_get_value(member, column_mapping.get('email', ''), ''),  # Email Address
            safe_get_value(member, column_mapping.get('accountability_buddies', ''), '')  # Group Mates
        ]
        ws.append(row)

        # Apply color coding and text formatting to the newly added row
        # Note: ws.append() creates cells, so we can format them after
        current_row = ws.max_row

        # Get the cells and apply formatting
        user_id_cell = ws.cell(row=current_row, column=1)
        name_cell = ws.cell(row=current_row, column=2)
        location_cell = ws.cell(row=current_row, column=3)

        # Ensure cells have values set (should already be set by ws.append)
        user_id_cell.value = member.get(column_mapping.get('user_id'), '')
        name_cell.value = formatted_name
        location_cell.value = location_display

        # Apply formatting to User ID, Name, and Location columns
        apply_color_to_cell(user_id_cell, sex, gender_identity, gender_preference, has_accountability_buddies, current_goal, is_user_id=True)  # User ID
        apply_color_to_cell(name_cell, sex, gender_identity, gender_preference, has_accountability_buddies, current_goal, is_user_id=False)  # Name
        apply_color_to_cell(location_cell, sex, gender_identity, gender_preference, has_accountability_buddies, current_goal, is_user_id=False)  # Location

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width


    # Check if filename_or_buffer is a string (file path) or BytesIO buffer
    if isinstance(filename_or_buffer, str):
        wb.save(filename_or_buffer)
    else:
        # It's a BytesIO buffer
        wb.save(filename_or_buffer)

def main():
    """Main function to read data and generate user list Excel"""
    # Read the merged Excel file
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name='Merged Data')
        print(f"✅ Successfully read input file with {len(df)} records")
    except Exception as e:
        try:
            df = pd.read_csv(INPUT_FILE)
            print(f"✅ Successfully read CSV file with {len(df)} records")
        except Exception as e2:
            print(f"❌ Error reading input file: {e2}")
            return

    # Find column mapping
    column_mapping = find_column_mapping(df)
    print("📋 Column mapping found:")
    for key, value in column_mapping.items():
        if value:
            print(f"  ✅ {key}: {value}")
        else:
            print(f"  ❌ {key}: NOT FOUND")

    # Convert DataFrame to list of dictionaries
    data = df.to_dict('records')

    # Remove duplicates based on user_id if available
    user_id_col = column_mapping.get('user_id')
    if user_id_col:
        unique_data = []
        seen_user_ids = set()

        for record in data:
            user_id_raw = record.get(user_id_col, '')
            user_id = str(user_id_raw).strip() if user_id_raw is not None else ''
            if user_id and user_id not in seen_user_ids:
                unique_data.append(record)
                seen_user_ids.add(user_id)
            elif not user_id:  # Keep records without user_id
                unique_data.append(record)

        data = unique_data
        print(f"📊 After deduplication: {len(data)} unique users (removed {len(df) - len(data)} duplicates)")
    else:
        print("⚠️ No user_id column found for deduplication")

    print(f"\n📊 Processing {len(data)} users...")

    # Save user list to Excel
    print("💾 Saving user list to Excel...")
    save_user_list_to_excel(data, OUTPUT_FILE, column_mapping)

    print("✅ User list generation completed successfully!")
    print(f"📁 Results saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
