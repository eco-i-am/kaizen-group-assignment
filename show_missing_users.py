import pandas as pd
import openpyxl
from collections import defaultdict

def show_missing_users():
    """Show detailed analysis of missing users by comparing input and output files"""
    
    input_file = 'sample_merged_data.xlsx'
    output_file = 'grouped_participants.xlsx'
    
    print("🔍 ANALYZING MISSING USERS")
    print("=" * 50)
    
    # Read input data
    try:
        df_input = pd.read_excel(input_file, sheet_name='Merged Data')
        print(f"✅ Input file: {len(df_input)} records")
    except Exception as e:
        print(f"❌ Error reading input file: {e}")
        return
    
    # Read output file
    try:
        wb_output = openpyxl.load_workbook(output_file)
        ws_output = wb_output.active
        print(f"✅ Output file: {ws_output.max_row - 1} rows (excluding header)")
    except Exception as e:
        print(f"❌ Error reading output file: {e}")
        return
    
    # Extract user IDs from input
    input_user_ids = set()
    input_user_details = {}
    
    for _, row in df_input.iterrows():
        user_id = row.get('user_id', '')
        if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
            user_id_str = str(user_id).strip()
            input_user_ids.add(user_id_str)
            input_user_details[user_id_str] = {
                'name': row.get('full_name', 'Unknown'),
                'gender_identity': row.get('gender_identity', 'Unknown'),
                'country': row.get('country', 'Unknown'),
                'city': row.get('city', 'Unknown'),
                'go_solo': row.get('prefer_solo', 'Unknown')
            }
    
    print(f"📊 Input user IDs: {len(input_user_ids)}")
    
    # Extract user IDs from output
    output_user_ids = set()
    output_user_groups = {}
    
    for row in range(2, ws_output.max_row + 1):  # Skip header
        group_name = ws_output.cell(row=row, column=1).value
        if not group_name:
            continue
        
        # Extract user IDs from columns 2, 5, 8, 11, 14, 17, 20 (User ID columns)
        for col in [2, 5, 8, 11, 14, 17, 20]:
            user_id = ws_output.cell(row=row, column=col).value
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                user_id_str = str(user_id).strip()
                output_user_ids.add(user_id_str)
                output_user_groups[user_id_str] = group_name
    
    print(f"📊 Output user IDs: {len(output_user_ids)}")
    
    # Find missing users
    missing_in_output = input_user_ids - output_user_ids
    extra_in_output = output_user_ids - input_user_ids
    
    print(f"\n📊 COMPARISON RESULTS:")
    print(f"  Users in input but missing from output: {len(missing_in_output)}")
    print(f"  Users in output but not in input: {len(extra_in_output)}")
    print(f"  Users in both: {len(input_user_ids & output_user_ids)}")
    
    # Show missing users in detail
    if missing_in_output:
        print(f"\n❌ MISSING USERS (in input but not in output):")
        print("-" * 50)
        
        for i, user_id in enumerate(sorted(missing_in_output, key=int), 1):
            details = input_user_details.get(user_id, {})
            print(f"{i:2d}. User ID: {user_id}")
            print(f"    Name: {details.get('name', 'Unknown')}")
            print(f"    Gender: {details.get('gender_identity', 'Unknown')}")
            print(f"    Country: {details.get('country', 'Unknown')}")
            print(f"    City: {details.get('city', 'Unknown')}")
            print(f"    Go Solo: {details.get('go_solo', 'Unknown')}")
            print()
    else:
        print(f"\n✅ NO MISSING USERS FOUND!")
        print("All users from the input file are present in the output file.")
    
    # Show extra users (if any)
    if extra_in_output:
        print(f"\n⚠️  EXTRA USERS (in output but not in input):")
        print("-" * 50)
        
        for i, user_id in enumerate(sorted(extra_in_output, key=int), 1):
            group = output_user_groups.get(user_id, 'Unknown Group')
            print(f"{i:2d}. User ID: {user_id} -> Group: {group}")
        print()
    
    # Show user distribution by group type
    print(f"\n📋 USER DISTRIBUTION BY GROUP TYPE:")
    print("-" * 50)
    
    group_types = defaultdict(list)
    for user_id, group_name in output_user_groups.items():
        if 'Solo' in str(group_name):
            group_types['Solo'].append(user_id)
        elif 'Requested Group' in str(group_name):
            group_types['Requested'].append(user_id)
        elif 'Team Group' in str(group_name):
            group_types['Team'].append(user_id)
        elif 'Excluded' in str(group_name):
            group_types['Excluded'].append(user_id)
        else:
            group_types['Regular'].append(user_id)
    
    for group_type, user_ids in group_types.items():
        print(f"{group_type:10}: {len(user_ids):2d} users")
        if len(user_ids) <= 10:
            print(f"           IDs: {sorted(user_ids, key=int)}")
        else:
            print(f"           IDs: {sorted(user_ids[:10], key=int)}... and {len(user_ids)-10} more")
        print()
    
    # Show some example groups
    print(f"\n📋 EXAMPLE GROUPS FROM OUTPUT:")
    print("-" * 50)
    
    count = 0
    for row in range(2, ws_output.max_row + 1):
        if count >= 10:  # Show first 10 groups
            break
        group_name = ws_output.cell(row=row, column=1).value
        if group_name:
            group_user_ids = []
            for col in [2, 5, 8, 11, 14, 17, 20]:
                user_id = ws_output.cell(row=row, column=col).value
                if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                    group_user_ids.append(str(user_id).strip())
            
            print(f"{count+1:2d}. {group_name}")
            print(f"    Users: {group_user_ids}")
            print()
            count += 1
    
    # Final summary
    print(f"\n📊 FINAL SUMMARY:")
    print("-" * 50)
    print(f"Total input users: {len(input_user_ids)}")
    print(f"Total output users: {len(output_user_ids)}")
    print(f"Missing users: {len(missing_in_output)}")
    print(f"Extra users: {len(extra_in_output)}")
    
    if len(missing_in_output) == 0:
        print(f"\n✅ SUCCESS: All users are accounted for!")
    else:
        print(f"\n❌ ISSUE: {len(missing_in_output)} users are missing from the output!")

if __name__ == "__main__":
    show_missing_users() 