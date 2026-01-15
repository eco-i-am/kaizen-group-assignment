"""
GROUP ASSIGNMENT TO EXCEL - Kaizen Participant Grouping System

This script processes merged participant data from Lazy Lifter API and creates
optimized group assignments based on accountability and geographic criteria:

1. ACCOUNTABILITY BUDDIES (Highest Priority)
   - Users who explicitly requested to be grouped with specific buddies
   - Uses graph-based clustering to find connected components
   - Groups users who reference each other as accountability partners

2. SOLO PARTICIPANTS (User Choice)
   - Users who prefer to work individually
   - Placed in single-member groups

3. REGULAR GROUPING (Algorithm-Based)
   - Hierarchical grouping by gender preference, then geography
   - Philippines: Province → City → Same-city groups
   - International: Country → State → Timezone regions
   - Optimizes group sizes (3-5 members)

4. SMALL GROUP MERGING (Optimization)
   - Merges groups <4 members based on geographic proximity
   - Ensures no group exceeds 5 members

FEATURES:
- Dynamic column mapping (handles various column name variations)
- Email normalization and alias mapping
- Geographic intelligence (timezones, Philippine regions)
- Gender identity respect (LGBTQ+ handling)
- Comprehensive diagnostic reporting
- Excel output with color coding and formatting

INPUT: Merged Excel file with user data and grouping preferences
OUTPUT: Excel file with organized groups, solos, and excluded users
"""

import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import re

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def safe_get_value(data_dict, key, default=''):
    """
    Safely retrieve a value from dictionary, handling NaN and None values.

    Args:
        data_dict: Dictionary containing participant data
        key: Column key to retrieve
        default: Default value if key not found or value is invalid

    Returns:
        str: Cleaned string value, or default if invalid
    """
    if not data_dict or key not in data_dict:
        return default
    value = data_dict[key]
    if pd.isna(value) or value is None:
        return default
    return str(value).strip()

# ============================================================================
# CONFIGURATION
# ============================================================================

# File paths - Update INPUT_FILE to point to your merged data file
INPUT_FILE = 'merged_users_grouping_preferences_20250719_133755.xlsx'  # Change this to your merged file
OUTPUT_FILE = 'grouped_participants.xlsx'

# ============================================================================
# EMAIL PROCESSING FUNCTIONS
# ============================================================================

def create_email_mapping(data, column_mapping):
    """
    Create email normalization mapping for known email aliases.

    Currently handles specific known cases where users have multiple email addresses
    or where accountability buddy references use different email formats.

    Args:
        data: Participant data (for future expansion)
        column_mapping: Column mappings (for future expansion)

    Returns:
        dict: Email mapping dictionary {old_email: canonical_email}
    """
    email_mapping = {}

    # Known email aliases - expand this as new cases are discovered
    # jaw.ybanez@yahoo.com is referenced in buddies but actual email is yo21st@gmail.com
    email_mapping['jaw.ybanez@yahoo.com'] = 'yo21st@gmail.com'

    return email_mapping

def check_name_similarity(name1, name2):
    """Check similarity between two names (simple implementation)"""
    if not name1 or not name2:
        return 0
    
    # Convert to lowercase and split into words
    words1 = set(str(name1).lower().split())
    words2 = set(str(name2).lower().split())
    
    # Calculate Jaccard similarity
    intersection = len(words1.intersection(words2))
    union = len(words1.union(words2))
    
    if union == 0:
        return 0
    
    return intersection / union

def normalize_email(email, email_mapping):
    """Normalize email addresses using the dynamic mapping"""
    if not email:
        return email
    
    email_lower = str(email).lower().strip()
    return email_mapping.get(email_lower, email_lower)

def extract_emails_from_accountability_buddies(accountability_buddies, email_mapping):
    """
    Extract emails from accountability_buddies field, handling multiple formats:

    SUPPORTED FORMATS:
    - Dictionary: {'1': 'email1', '2': 'email2'}
    - List with names: ['Name (email@domain.com)', 'Another Name (email2@domain.com)']
    - List with emails: ['email1@domain.com', 'email2@domain.com']
    - String representation of lists/dicts (JSON-like)
    - Simple comma-separated strings

    Args:
        accountability_buddies: The accountability buddies data in any supported format
        email_mapping: Dictionary for email normalization/alias mapping

    Returns:
        list: List of normalized email addresses
    """
    if not accountability_buddies:
        return []
    
    if isinstance(accountability_buddies, str):
        # Try to parse as dictionary first (for string representations of dicts)
        try:
            import ast
            parsed_dict = ast.literal_eval(accountability_buddies)
            if isinstance(parsed_dict, dict):
                # Handle dictionary format with numbered keys like {'1': 'email1', '2': 'email2'}
                emails = []
                for value in parsed_dict.values():
                    if value and '@' in str(value):
                        emails.append(normalize_email(str(value).strip(), email_mapping))
                return emails
        except (ValueError, SyntaxError):
            # Not a valid dict string, continue with string parsing
            pass
        
        # Handle regular string format (list-like strings)
        # Remove brackets and quotes, split by comma
        cleaned = accountability_buddies.strip('[]').replace('"', '').replace("'", '')
        emails = []
        for email_item in cleaned.split(','):
            email_item = email_item.strip()
            if email_item:
                # Check for "Name (email)" format
                import re
                email_match = re.search(r'\(([^)]+@[^)]+)\)', email_item)
                if email_match:
                    # Extract email from parentheses
                    email = email_match.group(1).strip()
                    if email and '@' in email:
                        emails.append(normalize_email(email, email_mapping))
                elif '@' in email_item:
                    # Direct email format
                    emails.append(normalize_email(email_item, email_mapping))
        return emails
    elif isinstance(accountability_buddies, dict):
        # Handle dictionary format with numbered keys like {'1': 'email1', '2': 'email2'}
        emails = []
        for value in accountability_buddies.values():
            if value and '@' in str(value):
                emails.append(normalize_email(str(value).strip(), email_mapping))
        return emails
    elif isinstance(accountability_buddies, list):
        # Handle list format - supports two formats:
        # 1. "Name (email@domain.com)" - extract email from parentheses
        # 2. "email@domain.com" - use email directly
        emails = []
        for item in accountability_buddies:
            if item:
                item_str = str(item).strip()

                # Check for "Name (email)" format
                import re
                email_match = re.search(r'\(([^)]+@[^)]+)\)', item_str)
                if email_match:
                    # Extract email from parentheses
                    email = email_match.group(1).strip()
                    if email and '@' in email:
                        emails.append(normalize_email(email, email_mapping))
                elif '@' in item_str:
                    # Direct email format
                    emails.append(normalize_email(item_str, email_mapping))
        return emails
    else:
        return []

# ============================================================================
# COLUMN MAPPING CONFIGURATION
# ============================================================================

# Dynamic column mapping - handles various column name variations
# The script automatically detects which columns exist in the input data
# Each key maps to possible column name variations
EXPECTED_COLUMNS = {
    'user_id': ['user_id', 'id', 'userid', 'user id', 'id_y', 'id_x'],
    'name': ['name', 'full_name', 'fullname', 'first_name', 'last_name', 'firstName', 'lastName'],
    'gender_identity': ['gender_identity', 'gender', 'genderidentity', 'genderIdentity'],
    'sex': ['sex', 'biological_sex', 'biologicalsex'],
    'residing_ph': ['residing_ph', 'residing_in_philippines', 'philippines_resident', 'lingInPhilippineExperience', 'residingInPhilippines'],
    'gender_preference': ['gender_preference', 'grouping_preference', 'preference', 'genderPref', 'groupGenderPreference'],
    'country': ['country', 'nationality'],
    'province': ['province', 'state_province'],
    'city': ['city', 'municipality'],
    'state': ['state', 'region', 'stat'],
    'go_solo': ['go_solo', 'solo', 'prefer_solo', 'goSolo'],
    'joining_as_student': ['joining_as_student', 'joiningAsStudent', 'student', 'is_student'],
    'kaizen_client_type': ['kaizen_client_type', 'kaizenClientType', 'client_type', 'clientType'],
    'accountability_buddies': ['accountability_buddies', 'accountabilityBuddies', 'accountability_buddies', 'buddies'],
    'has_accountability_buddies': ['has_accountability_buddies', 'hasAccountabilityBuddies', 'has_buddies'],
    'email': ['email', 'user_email', 'email_address', 'useremail', 'userEmail'],
    'previous_coach_name': ['previousCoachName', 'previous_coach_name', 'prev_coach_name', 'coach_name'],
    'current_goal': ['currentGoal', 'current_goal', 'goal'],
    'age_group': ['ageGroup', 'age_group', 'age']
}

# ============================================================================
# VISUAL FORMATTING CONSTANTS
# ============================================================================

# Color coding for Excel output based on participant characteristics
SEX_COLOR = {
    'male': 'ADD8E6',    # Light Blue for male participants
    'female': 'FFC0CB',  # Pink for female participants
}

# Special font color for LGBTQ+ participants
LGBTQ_FONT_COLOR = '800000'  # Maroon font color

# Fill color for participants with 'get_bigger' goal
GREEN_COLOR = '90EE90'  # Light Green fill for User IDs

# ============================================================================
# LOCATION FORMATTING FUNCTIONS
# ============================================================================

def format_location_display(member, column_mapping):
    """
    Format location display based on whether participant resides in Philippines.

    PHILIPPINES FORMAT: "City, Province"
    INTERNATIONAL FORMAT: "City, State, Location Identifier, Country"

    Uses enhanced logic to prioritize international-specific columns when available.

    Args:
        member: Participant data dictionary
        column_mapping: Column name mappings

    Returns:
        str: Formatted location string
    """
    residing_ph = safe_get_value(member, column_mapping.get('residing_ph', ''), '0').lower()

    if residing_ph in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
        # Philippines resident - show "city, province" format
        city = safe_get_value(member, column_mapping.get('city', ''), '')
        province = safe_get_value(member, column_mapping.get('province', ''), '')

        # Use "MM" as acronym for Metro Manila
        if province and str(province).lower() == 'metro manila':
            province = 'MM'

        if city and province:
            return f"{city}, {province}"
        elif city:
            return city
        elif province:
            return province
        else:
            return ''
    else:
        # International resident - show enhanced format: international_city, international_state, location_identifier, country
        # Check if international-specific columns exist, otherwise fall back to regular columns
        international_city = member.get('internationalCity', '') or safe_get_value(member, column_mapping.get('city', ''), '')
        international_state = member.get('internationalState', '') or safe_get_value(member, column_mapping.get('state', ''), '')
        location_identifier = member.get('locationIdentifier', '') or safe_get_value(member, 'location_identifier', '')
        country = safe_get_value(member, column_mapping.get('country', ''), '')

        parts = []
        if international_city:
            parts.append(str(international_city))
        if international_state:
            parts.append(str(international_state))
        if location_identifier:
            parts.append(str(location_identifier))
        if country:
            parts.append(str(country))

        return ', '.join(parts) if parts else ''

# ============================================================================
# GEOGRAPHIC GROUPING CONFIGURATION
# ============================================================================

# Geographic regions for international participant grouping
# Used when participants prefer no geographic restrictions
SIMILAR_COUNTRIES = {
    'southeast_asia': ['Philippines', 'Indonesia', 'Malaysia', 'Thailand', 'Vietnam', 'Singapore', 'Myanmar', 'Cambodia', 'Laos', 'Brunei'],
    'east_asia': ['China', 'Japan', 'South Korea', 'Taiwan', 'Hong Kong', 'Macau'],
    'south_asia': ['India', 'Pakistan', 'Bangladesh', 'Sri Lanka', 'Nepal', 'Bhutan', 'Maldives'],
    'north_america': ['United States', 'Canada', 'Mexico'],
    'europe': ['United Kingdom', 'Germany', 'France', 'Italy', 'Spain', 'Netherlands', 'Belgium', 'Switzerland', 'Austria', 'Sweden', 'Norway', 'Denmark', 'Finland'],
    'middle_east': ['Saudi Arabia', 'United Arab Emirates', 'Qatar', 'Kuwait', 'Bahrain', 'Oman', 'Jordan', 'Lebanon', 'Israel', 'Turkey'],
    'africa': ['South Africa', 'Nigeria', 'Kenya', 'Egypt', 'Morocco', 'Ghana', 'Ethiopia'],
    'oceania': ['Australia', 'New Zealand', 'Fiji', 'Papua New Guinea']
}

# Philippine administrative regions and their provinces/cities
# Used for geographic grouping of Philippines-based participants
PHILIPPINES_REGIONS = {
    'luzon': [
        'Metro Manila', 'Manila', 'Quezon City', 'Caloocan', 'Las Piñas', 'Makati', 'Malabon', 'Mandaluyong', 'Marikina', 'Muntinlupa', 'Navotas', 'Parañaque', 'Pasay', 'Pasig', 'San Juan', 'Taguig', 'Valenzuela', 'Pateros',
        'Bataan', 'Bulacan', 'Nueva Ecija', 'Pampanga', 'Tarlac', 'Zambales', 'Aurora',
        'Batangas', 'Cavite', 'Laguna', 'Quezon', 'Rizal',
        'Albay', 'Camarines Norte', 'Camarines Sur', 'Catanduanes', 'Masbate', 'Sorsogon',
        'Abra', 'Apayao', 'Benguet', 'Ifugao', 'Kalinga', 'Mountain Province',
        'Ilocos Norte', 'Ilocos Sur', 'La Union', 'Pangasinan',
        'Batanes', 'Cagayan', 'Isabela', 'Nueva Vizcaya', 'Quirino'
    ],
    'visayas': [
        'Aklan', 'Antique', 'Capiz', 'Iloilo', 'Guimaras', 'Negros Occidental',
        'Bohol', 'Cebu', 'Negros Oriental', 'Siquijor',
        'Biliran', 'Eastern Samar', 'Leyte', 'Northern Samar', 'Samar', 'Southern Leyte'
    ],
    'mindanao': [
        'Bukidnon', 'Camiguin', 'Lanao del Norte', 'Misamis Occidental', 'Misamis Oriental',
        'Davao del Norte', 'Davao del Sur', 'Davao Occidental', 'Davao Oriental', 'Davao de Oro',
        'Cotabato', 'Sarangani', 'South Cotabato', 'Sultan Kudarat',
        'Agusan del Norte', 'Agusan del Sur', 'Dinagat Islands', 'Surigao del Norte', 'Surigao del Sur',
        'Basilan', 'Lanao del Sur', 'Maguindanao', 'Sulu', 'Tawi-Tawi',
        'Zamboanga del Norte', 'Zamboanga del Sur', 'Zamboanga Sibugay'
    ]
}

def get_philippines_region(province):
    """Get the Philippines region for a given province"""
    if not province:
        return 'unknown'
    
    province_lower = str(province).strip().lower()
    
    for region, provinces in PHILIPPINES_REGIONS.items():
        for region_province in provinces:
            if province_lower == region_province.lower():
                return region
    
    # If not found in the mapping, try to guess based on common patterns
    if any(keyword in province_lower for keyword in ['manila', 'quezon', 'caloocan', 'makati', 'pasig', 'taguig', 'marikina', 'mandaluyong', 'las piñas', 'parañaque', 'muntinlupa', 'valenzuela', 'malabon', 'navotas', 'san juan', 'pasay', 'pateros']):
        return 'luzon'
    elif any(keyword in province_lower for keyword in ['cebu', 'iloilo', 'bohol', 'negros', 'samar', 'leyte', 'aklan', 'antique', 'capiz', 'guimaras', 'siquijor', 'biliran']):
        return 'visayas'
    elif any(keyword in province_lower for keyword in ['davao', 'cotabato', 'zamboanga', 'bukidnon', 'misamis', 'agusan', 'surigao', 'lanao', 'maguindanao', 'sulu', 'tawi-tawi', 'basilan']):
        return 'mindanao'
    
    return 'unknown'

# Define timezone regions for international grouping
TIMEZONE_REGIONS = {
    'pst_pdt': ['United States', 'Canada'],  # Pacific Time
    'mst_mdt': ['United States', 'Canada'],  # Mountain Time
    'cst_cdt': ['United States', 'Canada'],  # Central Time
    'est_edt': ['United States', 'Canada'],  # Eastern Time
    'gmt_bst': ['United Kingdom', 'Ireland', 'Portugal','Isle of Man'],  # GMT/BST
    'cet_cest': ['Germany', 'France', 'Italy', 'Spain', 'Netherlands', 'Belgium', 'Switzerland', 'Austria', 'Sweden', 'Norway', 'Denmark', 'Finland', 'Poland', 'Czech Republic', 'Hungary', 'Slovakia', 'Slovenia', 'Croatia', 'Serbia', 'Bosnia', 'Montenegro', 'North Macedonia', 'Albania', 'Kosovo', 'Bulgaria', 'Romania', 'Moldova', 'Ukraine', 'Belarus', 'Lithuania', 'Latvia', 'Estonia'],
    'eet_eest': ['Greece', 'Cyprus', 'Bulgaria', 'Romania', 'Moldova', 'Ukraine', 'Belarus', 'Lithuania', 'Latvia', 'Estonia', 'Finland'],
    'msk': ['Russia'],  # Moscow Time
    'ist': ['India', 'Sri Lanka'],  # India Standard Time
    'pkt': ['Pakistan'],  # Pakistan Time
    'bst': ['Bangladesh'],  # Bangladesh Time
    'jst': ['Japan', 'South Korea'],  # Japan Standard Time
    'cst': ['China', 'Taiwan', 'Hong Kong', 'Macau'],  # China Standard Time
    'sgt': ['Singapore', 'Malaysia', 'Brunei'],  # Singapore Time
    'ict': ['Thailand', 'Vietnam', 'Cambodia', 'Laos'],  # Indochina Time
    'wib': ['Indonesia'],  # Western Indonesian Time
    'aest_aedt': ['Australia'],  # Australian Eastern Time
    'nzst_nzdt': ['New Zealand'],  # New Zealand Time
    'gst': ['United Arab Emirates', 'Oman'],  # Gulf Standard Time
    'ast': ['Saudi Arabia', 'Kuwait', 'Bahrain', 'Qatar'],  # Arabia Standard Time
    'eat': ['Kenya', 'Ethiopia', 'Tanzania', 'Uganda', 'Rwanda', 'Burundi', 'Somalia', 'Djibouti', 'Eritrea'],  # East Africa Time
    'wast_wat': ['Nigeria', 'Ghana', 'Cameroon', 'Chad', 'Central African Republic', 'Gabon', 'Congo', 'DR Congo', 'Angola'],  # West Africa Time
    'sast': ['South Africa', 'Namibia', 'Botswana', 'Zimbabwe', 'Zambia', 'Malawi', 'Mozambique', 'Lesotho', 'Eswatini'],  # South Africa Time
    'est': ['Egypt', 'Libya', 'Sudan', 'South Sudan'],  # Egypt Standard Time
    'pst': ['Mexico'],  # Pacific Standard Time (Mexico)
    'cst': ['Mexico'],  # Central Standard Time (Mexico)
    'est': ['Mexico'],  # Eastern Standard Time (Mexico)
    'cayman_est': ['Cayman Islands'],
    'bermuda_ast': ['Bermuda'],
    # Geographic regions
    'southeast_asia': 'GMT+7',
    'east_asia': 'GMT+8',
    'south_asia': 'GMT+5',   
    'north_america': 'GMT-5', 
    'europe': 'GMT+1',
    'middle_east': 'GMT+3', 
    'africa': 'GMT+2',
    'oceania': 'GMT+10'
}

# GMT offset mapping for timezone display labels
# Converts timezone region codes to human-readable GMT offset labels
GMT_OFFSETS = {
    'pst_pdt': 'GMT-8',
    'mst_mdt': 'GMT-7', 
    'cst_cdt': 'GMT-6',
    'est_edt': 'GMT-5',
    'gmt_bst': 'GMT+0',
    'cet_cest': 'GMT+1',
    'eet_eest': 'GMT+2',
    'msk': 'GMT+3',
    'ist': 'GMT+5:30',
    'pkt': 'GMT+5',
    'bst': 'GMT+6',
    'jst': 'GMT+9',
    'cst': 'GMT+8',
    'sgt': 'GMT+8',
    'ict': 'GMT+7',
    'wib': 'GMT+7',
    'aest_aedt': 'GMT+10',
    'nzst_nzdt': 'GMT+12',
    'gst': 'GMT+4',
    'ast': 'GMT+3',
    'eat': 'GMT+3',
    'wast_wat': 'GMT+1',
    'sast': 'GMT+2',
    'est': 'GMT+2',
    'pst': 'GMT-8',
    'cst': 'GMT-6',
    'est': 'GMT-5',
    'cayman_est': 'GMT-5',
    'bermuda_ast': 'GMT-4',
    # Geographic regions
    'southeast_asia': 'GMT+7',
    'east_asia': 'GMT+8',
    'south_asia': 'GMT+5',   
    'north_america': 'GMT-5', 
    'europe': 'GMT+1',
    'middle_east': 'GMT+3', 
    'africa': 'GMT+2',
    'oceania': 'GMT+10'
}

def get_gmt_offset_value(timezone_region):
    """Get GMT offset as a numeric value for sorting (negative for behind GMT, positive for ahead)"""
    if timezone_region in GMT_OFFSETS:
        offset_str = GMT_OFFSETS[timezone_region]
        # Extract numeric value from GMT+X or GMT-X or GMT+X:Y
        if 'GMT+' in offset_str:
            # Handle half-hour offsets like GMT+5:30
            parts = offset_str.replace('GMT+', '').split(':')
            hours = int(parts[0])
            minutes = int(parts[1]) if len(parts) > 1 else 0
            return hours + (minutes / 60.0)
        else:
            # Handle whole hour offsets
            return int(offset_str.replace('GMT+', '').replace('GMT-', '-'))
    return 0  # Default for unknown timezones

def normalize_country_name(country):
    if not country:
        return ''
    country = str(country).strip().lower().replace('.', '')
    country = re.sub(r'\s+', ' ', country)  # collapse multiple spaces
    return country

def extract_country_from_field(field_value):
    if not field_value:
        return field_value
    field_str = str(field_value).strip()
    if field_str.lower() in ['nan', 'none', '[]']:
        return field_str
    if ',' in field_str:
        parts = [part.strip() for part in field_str.split(',')]
        return parts[-1]
    # Handle abbreviations and variations
    country_mappings = {
        'uae': 'united arab emirates',
        'u a e': 'united arab emirates',
        'u.a.e.': 'united arab emirates',
        'united arab emirates': 'united arab emirates',
        'usa': 'united states',
        'us': 'united states',
        'u s a': 'united states',
        'u.s.a.': 'united states',
        'uk': 'united kingdom',
        'u k': 'united kingdom',
        'u.k.': 'united kingdom',
        'north ame': 'united states',
        'north america': 'united states',
        'bermuda': 'bermuda',
        'cayman islands': 'cayman islands',
    }
    norm = normalize_country_name(field_str)
    if norm in country_mappings:
        return country_mappings[norm]
    return field_str

def get_timezone_region(country, state=None):
    original_country = country
    country = extract_country_from_field(country)
    norm_country = normalize_country_name(country)
    # Special handling for US states with different timezones
    if norm_country == 'united states' and state:
        state = str(state).strip().lower()
        if state in ['california', 'washington', 'oregon', 'nevada', 'alaska']:
            return 'pst_pdt'
        elif state in ['colorado', 'utah', 'wyoming', 'montana', 'idaho', 'new mexico', 'arizona']:
            return 'mst_mdt'
        elif state in ['texas', 'oklahoma', 'kansas', 'nebraska', 'south dakota', 'north dakota', 'minnesota', 'iowa', 'missouri', 'arkansas', 'louisiana', 'mississippi', 'alabama', 'illinois', 'wisconsin', 'michigan', 'indiana', 'kentucky', 'tennessee']:
            return 'cst_cdt'
        elif state in ['new york', 'new jersey', 'pennsylvania', 'ohio', 'indiana', 'michigan', 'illinois', 'wisconsin', 'minnesota', 'iowa', 'missouri', 'arkansas', 'louisiana', 'mississippi', 'alabama', 'georgia', 'florida', 'south carolina', 'north carolina', 'virginia', 'west virginia', 'maryland', 'delaware', 'new hampshire', 'vermont', 'maine', 'massachusetts', 'rhode island', 'connecticut']:
            return 'est_edt'
    # Check timezone regions with case-insensitive matching
    for timezone, countries in TIMEZONE_REGIONS.items():
        for mapped_country in countries:
            if norm_country == normalize_country_name(mapped_country):
                return timezone
    # Fallback to geographic regions with case-insensitive matching
    for region, countries in SIMILAR_COUNTRIES.items():
        for mapped_country in countries:
            if norm_country == normalize_country_name(mapped_country):
                return region
    return 'other'

def get_timezone_label(timezone_region):
    """Get timezone label with GMT offset first"""
    if timezone_region in GMT_OFFSETS:
        return f"{GMT_OFFSETS[timezone_region]} => {timezone_region.upper()}"
    elif timezone_region in SIMILAR_COUNTRIES:
        return f"{timezone_region.replace('_', ' ').title()}"
    else:
        return timezone_region

def find_column_mapping(df):
    """Dynamically find column mapping based on available columns"""
    mapping = {}
    available_columns = [col.lower() for col in df.columns]
    
    for expected_key, possible_names in EXPECTED_COLUMNS.items():
        found = False
        for possible_name in possible_names:
            if possible_name.lower() in available_columns:
                # Find the exact column name (case-insensitive match)
                for col in df.columns:
                    if col.lower() == possible_name.lower():
                        mapping[expected_key] = col
                        found = True
                        break
                if found:
                    break
        
        if not found:
            mapping[expected_key] = None
    
    return mapping

def get_country_region(country):
    """Get the region for a given country"""
    country = str(country).strip()
    for region, countries in SIMILAR_COUNTRIES.items():
        if country in countries:
            return region
    return 'other'

def apply_color_to_cell(cell, sex, gender_identity=None, gender_preference=None, has_accountability_buddies=None, current_goal=None, is_user_id=False):
    """Apply color coding based on sex, font styling, and special fill coloring"""
    # Apply fill color based on sex (default)
    fill_color = None
    sex_lower = str(sex).lower().strip() if sex else ''
    if sex_lower in SEX_COLOR:
        fill_color = SEX_COLOR[sex_lower]

    # Special fill color for get_bigger goal (overrides sex color for User ID)
    if is_user_id and current_goal and str(current_goal).lower() == 'get_bigger':
        fill_color = GREEN_COLOR

    # Apply fill color if set
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Apply font formatting
    font_color = None
    is_bold = False
    is_underline = False

    # Maroon font color for LGBTQ+
    if gender_identity and str(gender_identity).lower() in ['lgbtq+', 'lgbtq']:
        font_color = LGBTQ_FONT_COLOR

    # Bold text for same_gender preference
    if gender_preference and str(gender_preference).lower().strip() == 'same_gender':
        is_bold = True

    # Underlined text for users with accountability buddies
    if has_accountability_buddies and str(has_accountability_buddies).lower() in ['1', '1.0', 'true', 'yes']:
        is_underline = True

    # Create font style - always create a new font with the desired properties
    cell.font = Font(
        color=font_color,
        bold=is_bold,
        underline='single' if is_underline else None
    )

def generate_diagnostic_report(user_tracking, original_count, solo_groups, grouped, excluded_users, requested_groups, column_mapping):
    """Generate a comprehensive diagnostic report of user distribution"""
    
    print(f"\n" + "="*60)
    print(f"📊 USER DISTRIBUTION DIAGNOSTIC REPORT")
    print(f"="*60)
    
    # Summary statistics
    print(f"\n📈 SUMMARY STATISTICS:")
    print(f"  Total original users: {original_count}")
    print(f"  Solo groups created: {len(solo_groups)}")
    print(f"  Regular groups created: {len(grouped)}")
    print(f"  Requested groups created: {len(requested_groups)}")
    print(f"  Excluded users: {len(excluded_users)}")
    
    # User status breakdown
    status_counts = {}
    for user_id, info in user_tracking.items():
        status = info['status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    print(f"\n👥 USER STATUS BREAKDOWN:")
    for status, count in status_counts.items():
        print(f"  {status.replace('_', ' ').title()}: {count} users")
    
    # Check for any unaccounted users
    total_accounted = sum(status_counts.values())
    missing_count = original_count - total_accounted
    
    if missing_count > 0:
        print(f"\n⚠️  MISSING USERS DETECTED:")
        print(f"  Missing/Unaccounted users: {missing_count}")
        print(f"  This indicates a potential issue in the grouping logic")
    else:
        print(f"\n✅ ALL USERS ACCOUNTED FOR:")
        print(f"  All {original_count} users have been properly categorized")
    
    # Detailed breakdown by status
    print(f"\n📋 DETAILED BREAKDOWN:")
    
    for status in ['excluded', 'accountability_buddies', 'solo', 'regular_grouping']:
        users_in_status = [u for u in user_tracking.values() if u['status'] == status]
        if users_in_status:
            print(f"\n  {status.upper().replace('_', ' ')} ({len(users_in_status)} users):")
            for user in users_in_status[:5]:  # Show first 5
                user_id = next(k for k, v in user_tracking.items() if v == user)
                print(f"    - {user_id}: {user['reason']}")
            if len(users_in_status) > 5:
                print(f"    ... and {len(users_in_status) - 5} more")
    
    # Check for potential issues
    print(f"\n🔍 POTENTIAL ISSUES:")
    
    # Check for users without emails
    users_without_email = [u for u in user_tracking.values() if not u['email'] or '@' not in u['email']]
    if users_without_email:
        print(f"  ⚠️  Users without valid emails: {len(users_without_email)}")
        print(f"    This may affect accountability buddy functionality")
    
    # Check for users with missing critical data
    critical_columns = ['gender_identity', 'gender_preference', 'residing_ph']
    missing_data_issues = []
    
    for col in critical_columns:
        missing_count = 0
        for user_id, info in user_tracking.items():
            row_data = info['row_data']
            # Get value using column mapping
            if isinstance(row_data, dict) and column_mapping and col in column_mapping:
                col_name = column_mapping[col]
                value = row_data.get(col_name, '') if col_name else ''
            else:
                value = ''
            if not value or str(value).strip() in ['', 'None', 'nan']:
                missing_count += 1
        
        if missing_count > 0:
            missing_data_issues.append(f"{col}: {missing_count} users")
    
    if missing_data_issues:
        print(f"  ⚠️  Users with missing critical data:")
        for issue in missing_data_issues:
            print(f"    - {issue}")
    
    # Group size analysis
    print(f"\n📊 GROUP SIZE ANALYSIS:")
    
    # Solo groups
    if solo_groups:
        solo_sizes = [len(group) for group in solo_groups]
        print(f"  Solo groups: {len(solo_groups)} groups, sizes: {solo_sizes}")
    
    # Regular groups
    if grouped:
        regular_sizes = [len(members) for members in grouped.values()]
        print(f"  Regular groups: {len(grouped)} groups")
        print(f"    Size distribution: {sorted(regular_sizes)}")
        print(f"    Average size: {sum(regular_sizes)/len(regular_sizes):.1f}")
        print(f"    Min size: {min(regular_sizes)}, Max size: {max(regular_sizes)}")
    
    # Requested groups
    if requested_groups:
        requested_sizes = [len(group) for group in requested_groups]
        print(f"  Requested groups: {len(requested_groups)} groups, sizes: {requested_sizes}")
    
    # Final verification
    print(f"\n✅ FINAL VERIFICATION:")
    print(f"  Input users: {original_count}")
    print(f"  Output users: {total_accounted}")
    print(f"  Status: {'✅ All users accounted for' if missing_count == 0 else '❌ Missing users detected'}")
    
    print(f"\n" + "="*60)
    print(f"📊 END OF DIAGNOSTIC REPORT")
    print(f"="*60)

def generate_missing_users_analysis(user_tracking, original_count, solo_groups, grouped, excluded_users, requested_groups, column_mapping):
    """Generate detailed analysis of missing users and user distribution"""
    
    print(f"\n" + "="*60)
    print(f"🔍 MISSING USERS ANALYSIS")
    print(f"="*60)
    
    # Collect all users from different sources
    all_tracked_users = set(user_tracking.keys())
    
    # Collect users from solo groups
    solo_users = set()
    for group in solo_groups:
        for member in group:
            if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                col_name = column_mapping['user_id']
                user_id = member.get(col_name, 'Unknown')
            else:
                user_id = 'Unknown'
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                solo_users.add(str(user_id).strip())
    
    # Collect users from regular groups
    regular_users = set()
    for group_name, members in grouped.items():
        for member in members:
            if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                col_name = column_mapping['user_id']
                user_id = member.get(col_name, 'Unknown')
            else:
                user_id = 'Unknown'
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                regular_users.add(str(user_id).strip())
    
    # Collect users from requested groups
    requested_users = set()
    for group in requested_groups:
        for member in group:
            if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                col_name = column_mapping['user_id']
                user_id = member.get(col_name, 'Unknown')
            else:
                user_id = 'Unknown'
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                requested_users.add(str(user_id).strip())
    
    # Collect users from excluded users
    excluded_user_ids = set()
    for user in excluded_users:
        if isinstance(user, dict) and column_mapping and 'user_id' in column_mapping:
            col_name = column_mapping['user_id']
            user_id = user.get(col_name, 'Unknown')
        else:
            user_id = 'Unknown'
        if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
            excluded_user_ids.add(str(user_id).strip())
    
    # Combine all output users
    all_output_users = solo_users | regular_users | requested_users | excluded_user_ids
    
    # Find missing users
    missing_users = all_tracked_users - all_output_users
    extra_users = all_output_users - all_tracked_users
    
    print(f"\n📊 USER COMPARISON:")
    print(f"  Total tracked users: {len(all_tracked_users)}")
    print(f"  Solo users: {len(solo_users)}")
    print(f"  Regular group users: {len(regular_users)}")
    print(f"  Requested group users: {len(requested_users)}")
    print(f"  Excluded users: {len(excluded_user_ids)}")
    print(f"  Total output users: {len(all_output_users)}")
    
    print(f"\n📊 MISSING USERS ANALYSIS:")
    print(f"  Users missing from output: {len(missing_users)}")
    print(f"  Extra users in output: {len(extra_users)}")
    print(f"  Users in both: {len(all_tracked_users & all_output_users)}")
    
    # Show missing users in detail
    if missing_users:
        print(f"\n❌ MISSING USERS (tracked but not in output):")
        print("-" * 50)
        
        for i, user_id in enumerate(sorted(missing_users, key=lambda x: int(x) if str(x).isdigit() else 999), 1):
            user_info = user_tracking.get(user_id, {})
            print(f"{i:2d}. User ID: {user_id}")
            print(f"    Status: {user_info.get('status', 'Unknown')}")
            print(f"    Reason: {user_info.get('reason', 'Unknown')}")
            print(f"    Email: {user_info.get('email', 'No email')}")
            print()
    else:
        print(f"\n✅ NO MISSING USERS FOUND!")
        print("All tracked users are present in the output groups.")
    
    # Show extra users (if any)
    if extra_users:
        print(f"\n⚠️  EXTRA USERS (in output but not tracked):")
        print("-" * 50)
        
        for i, user_id in enumerate(sorted(extra_users, key=lambda x: int(x) if str(x).isdigit() else 999), 1):
            print(f"{i:2d}. User ID: {user_id}")
        print()
    
    # Show user distribution by group type
    print(f"\n📋 USER DISTRIBUTION BY GROUP TYPE:")
    print("-" * 50)
    
    group_types = {
        'Solo': solo_users,
        'Regular': regular_users,
        'Requested': requested_users,
        'Excluded': excluded_user_ids
    }
    
    for group_type, user_ids in group_types.items():
        if user_ids:
            print(f"{group_type:10}: {len(user_ids):2d} users")
            sorted_ids = sorted(user_ids, key=lambda x: int(x) if str(x).isdigit() else 999)
            if len(sorted_ids) <= 10:
                print(f"           IDs: {sorted_ids}")
            else:
                print(f"           IDs: {sorted_ids[:10]}... and {len(sorted_ids)-10} more")
            print()
    
    # Show some example groups
    print(f"\n📋 EXAMPLE GROUPS:")
    print("-" * 50)
    
    # Solo groups
    if solo_groups:
        print("Solo Groups:")
        for i, group in enumerate(solo_groups[:5], 1):
            user_ids = []
            for member in group:
                if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                    col_name = column_mapping['user_id']
                    user_id = member.get(col_name, 'Unknown')
                else:
                    user_id = 'Unknown'
                if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                    user_ids.append(str(user_id).strip())
            print(f"  {i}. Solo {i}: {user_ids}")
        if len(solo_groups) > 5:
            print(f"  ... and {len(solo_groups) - 5} more solo groups")
        print()
    
    # Regular groups
    if grouped:
        print("Regular Groups (first 5):")
        count = 0
        for group_name, members in grouped.items():
            if count >= 5:
                break
            user_ids = []
            for member in members:
                if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                    col_name = column_mapping['user_id']
                    user_id = member.get(col_name, 'Unknown')
                else:
                    user_id = 'Unknown'
                if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                    user_ids.append(str(user_id).strip())
            print(f"  {count+1}. {group_name}: {user_ids}")
            count += 1
        if len(grouped) > 5:
            print(f"  ... and {len(grouped) - 5} more regular groups")
        print()
    
    # Requested groups
    if requested_groups:
        print("Requested Groups:")
        for i, group in enumerate(requested_groups[:5], 1):
            user_ids = []
            for member in group:
                if isinstance(member, dict) and column_mapping and 'user_id' in column_mapping:
                    col_name = column_mapping['user_id']
                    user_id = member.get(col_name, 'Unknown')
                else:
                    user_id = 'Unknown'
                if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                    user_ids.append(str(user_id).strip())
            print(f"  {i}. Requested Group {i}: {user_ids}")
        if len(requested_groups) > 5:
            print(f"  ... and {len(requested_groups) - 5} more requested groups")
        print()
    
    # Final summary
    print(f"\n📊 FINAL SUMMARY:")
    print("-" * 50)
    print(f"Total tracked users: {len(all_tracked_users)}")
    print(f"Total output users: {len(all_output_users)}")
    print(f"Missing users: {len(missing_users)}")
    print(f"Extra users: {len(extra_users)}")
    
    if len(missing_users) == 0:
        print(f"\n✅ SUCCESS: All users are accounted for!")
    else:
        print(f"\n❌ ISSUE: {len(missing_users)} users are missing from the output!")
    
    print(f"\n" + "="*60)
    print(f"🔍 END OF MISSING USERS ANALYSIS")
    print(f"="*60)

def check_for_duplicates(solo_groups, grouped, excluded_users, requested_groups, column_mapping):
    """Check for duplicate users across all group types and report them"""
    
    print(f"\n" + "="*60)
    print(f"🔍 DUPLICATE USERS DETECTION")
    print(f"="*60)
    
    all_users = set()
    duplicate_users = set()
    user_locations = {}  # Track where each user appears
    
    def get_value(row, key, default=''):
        if column_mapping and key in column_mapping:
            if isinstance(row, dict):
                return row.get(column_mapping[key], default)
            else:
                return default
        else:
            return default
    
    # Check solo groups
    for i, group in enumerate(solo_groups):
        for member in group:
            user_id = get_value(member, 'user_id', 'Unknown')
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                user_key = str(user_id).strip()
                if user_key in all_users:
                    duplicate_users.add(user_key)
                    if user_key not in user_locations:
                        user_locations[user_key] = []
                    user_locations[user_key].append(f"Solo Group {i+1}")
                else:
                    all_users.add(user_key)
                    user_locations[user_key] = [f"Solo Group {i+1}"]
    
    # Check regular groups
    for group_name, members in grouped.items():
        for member in members:
            user_id = get_value(member, 'user_id', 'Unknown')
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                user_key = str(user_id).strip()
                if user_key in all_users:
                    duplicate_users.add(user_key)
                    if user_key not in user_locations:
                        user_locations[user_key] = []
                    user_locations[user_key].append(f"Regular Group: {group_name}")
                else:
                    all_users.add(user_key)
                    user_locations[user_key] = [f"Regular Group: {group_name}"]
    
    # Check requested groups
    for i, group in enumerate(requested_groups):
        for member in group:
            user_id = get_value(member, 'user_id', 'Unknown')
            if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
                user_key = str(user_id).strip()
                if user_key in all_users:
                    duplicate_users.add(user_key)
                    if user_key not in user_locations:
                        user_locations[user_key] = []
                    user_locations[user_key].append(f"Requested Group {i+1}")
                else:
                    all_users.add(user_key)
                    user_locations[user_key] = [f"Requested Group {i+1}"]
    
    # Check excluded users
    for user in excluded_users:
        user_id = get_value(user, 'user_id', 'Unknown')
        if user_id and str(user_id).strip() not in ['', 'nan', 'None']:
            user_key = str(user_id).strip()
            if user_key in all_users:
                duplicate_users.add(user_key)
                if user_key not in user_locations:
                    user_locations[user_key] = []
                user_locations[user_key].append("Excluded Users")
            else:
                all_users.add(user_key)
                user_locations[user_key] = ["Excluded Users"]
    
    # Report results
    print(f"\n📊 DUPLICATE DETECTION RESULTS:")
    print(f"  Total unique users in output: {len(all_users)}")
    print(f"  Duplicate users found: {len(duplicate_users)}")
    
    if duplicate_users:
        print(f"\n❌ DUPLICATE USERS DETECTED:")
        print("-" * 60)
        for user_id in sorted(duplicate_users, key=int):
            locations = user_locations.get(user_id, [])
            print(f"  User ID: {user_id}")
            print(f"    Appears in: {', '.join(locations)}")
            print()
        
        print(f"⚠️  WARNING: {len(duplicate_users)} users appear in multiple groups!")
        print(f"   This indicates a logic error in the grouping process.")
    else:
        print(f"\n✅ NO DUPLICATE USERS FOUND!")
        print(f"   All users appear in exactly one group type.")
    
    print(f"\n" + "="*60)
    print(f"🔍 END OF DUPLICATE USERS DETECTION")
    print(f"="*60)

def merge_small_groups(grouped, column_mapping):
    """
    POST-PROCESSING OPTIMIZATION: Merge small groups to improve group sizes.

    Identifies regular groups with <4 members and merges them based on geographic
    proximity and gender compatibility. Ensures all groups have optimal sizes
    (3-5 members) while maintaining geographic and gender preferences.

    MERGING STRATEGY:
    1. Find small groups (<4 members) in regular algorithmic groups
    2. Group small groups by geographic location + gender compatibility
    3. Merge compatible groups, prioritizing same-location matches
    4. Never exceed 5 members per group

    Args:
        grouped: Dictionary of {group_name: [members]} from regular grouping
        column_mapping: Column name mappings

    Returns:
        dict: Optimized groups with improved size distribution
    """
    if not grouped:
        return grouped
    
    # Helper function to get value safely
    def get_value(row, key, default=''):
        if column_mapping and key in column_mapping:
            if isinstance(row, dict):
                return row.get(column_mapping[key], default)
            else:
                return default
        else:
            # Fallback to old format (list indices)
            if isinstance(row, list):
                if key == 'user_id':
                    return row[0] if len(row) > 0 else default
                elif key == 'residing_ph':
                    return row[8] if len(row) > 8 else default
                elif key == 'country':
                    return row[16] if len(row) > 16 else default
                elif key == 'province':
                    return row[17] if len(row) > 17 else default
                elif key == 'city':
                    return row[18] if len(row) > 18 else default
                elif key == 'state':
                    return row[19] if len(row) > 19 else default
            return default
    
    # Helper function to get location key for proximity matching (more flexible)
    def get_location_key(member):
        residing_ph = str(get_value(member, 'residing_ph', '0')).strip().lower()
        if residing_ph in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
            # Philippines: use province only (more flexible than exact city)
            province = str(get_value(member, 'province', '')).strip().lower()
            return f"PH_{province}"
        else:
            # International: use country and timezone region (more flexible than exact state)
            country = str(get_value(member, 'country', '')).strip().lower()
            state = str(get_value(member, 'state', '')).strip().lower()
            timezone_region = get_timezone_region(country, state)
            return f"INT_{country}_{timezone_region}"
    
    # Helper function to get gender key for compatibility
    def get_gender_key(member):
        gender_pref = str(get_value(member, 'gender_preference', '')).lower()
        if gender_pref == 'same_gender':
            sex = str(get_value(member, 'sex', '')).lower()
            gender_identity = str(get_value(member, 'gender_identity', '')).upper()
            if gender_identity == 'LGBTQ+':
                return f"lgbtq+_{sex}"
            else:
                return sex
        elif gender_pref == 'no_preference':
            return 'no_preference'
        else:
            return 'other'
    
    # Separate small groups (less than 4 members) that start with "Group*"
    small_groups = {}
    other_groups = {}
    
    for group_name, members in grouped.items():
        if group_name.startswith("Group ") and len(members) < 4:
            small_groups[group_name] = members
        else:
            other_groups[group_name] = members
    
    if not small_groups:
        return grouped
    
    print(f"🔍 Found {len(small_groups)} small groups to potentially merge")
    
    # Group small groups by location and gender compatibility
    location_gender_groups = defaultdict(list)
    
    for group_name, members in small_groups.items():
        if not members:
            continue
        
        # Get location and gender info from first member
        first_member = members[0]
        location_key = get_location_key(first_member)
        gender_key = get_gender_key(first_member)
        
        # Create a composite key for grouping
        composite_key = f"{location_key}_{gender_key}"
        location_gender_groups[composite_key].append((group_name, members))
    
    # If we still have groups with <4 members, try broader geographic matching
    # Group by just gender preference (ignore location constraints)
    if any(len(group_list) == 1 for group_list in location_gender_groups.values()):
        print("🔄 Trying broader geographic matching for remaining small groups...")
        
        # Reset and try broader matching
        location_gender_groups = defaultdict(list)
        
        for group_name, members in small_groups.items():
            if not members:
                continue
            
            first_member = members[0]
            gender_key = get_gender_key(first_member)
            
            # Use only gender key for broader matching
            location_gender_groups[gender_key].append((group_name, members))
    
    # Merge groups within each location-gender combination
    merged_groups = {}
    group_counter = 1
    
    for composite_key, group_list in location_gender_groups.items():
        if len(group_list) == 1:
            # Only one group in this location-gender combo, keep as is
            group_name, members = group_list[0]
            merged_groups[group_name] = members
            continue
        
        # Multiple groups in same location-gender combo, try to merge
        all_members = []
        for _, members in group_list:
            all_members.extend(members)
        
        # Create new groups of up to 5 members
        i = 0
        while i < len(all_members):
            group_members = all_members[i:i+5]
            
            # Create new group name with location info
            first_member = group_members[0]
            location_key = get_location_key(first_member)
            gender_key = get_gender_key(first_member)
            
            # Extract location info for display
            if location_key.startswith("PH_"):
                parts = location_key.split("_")
                if len(parts) >= 2:
                    province = parts[1].title()
                    location_info = f"Province: {province}"
                else:
                    location_info = "Philippines"
            else:
                parts = location_key.split("_")
                if len(parts) >= 3:
                    country = parts[1].title()
                    timezone_region = parts[2].title()
                    location_info = f"Country: {country}, Timezone: {timezone_region}"
                else:
                    location_info = "International"
            
            new_group_name = f"Group {group_counter} ({gender_key}, {location_info})"
            merged_groups[new_group_name] = group_members
            group_counter += 1
            i += 5
    
    # Final check: if any merged groups still have <4 members, combine them
    final_small_groups = {}
    final_regular_groups = {}
    
    for group_name, members in merged_groups.items():
        if len(members) < 4:
            final_small_groups[group_name] = members
        else:
            final_regular_groups[group_name] = members
    
    # If we still have small groups, combine them by gender preference only
    if final_small_groups:
        print("🔄 Final merge: combining remaining small groups by gender preference...")
        
        gender_groups = defaultdict(list)
        for group_name, members in final_small_groups.items():
            if not members:
                continue
            
            first_member = members[0]
            gender_key = get_gender_key(first_member)
            gender_groups[gender_key].extend(members)
        
        # Create final groups from each gender category
        for gender_key, all_members in gender_groups.items():
            i = 0
            while i < len(all_members):
                group_members = all_members[i:i+5]
                new_group_name = f"Group {group_counter} ({gender_key}, merged)"
                final_regular_groups[new_group_name] = group_members
                group_counter += 1
                i += 5
    
    # Combine all groups
    final_groups = {}
    final_groups.update(other_groups)
    final_groups.update(final_regular_groups)
    
    print(f"✅ Merged {len(small_groups)} small groups into {len(final_regular_groups)} groups")
    
    return final_groups

# ============================================================================
# MAIN GROUPING FUNCTIONS
# ============================================================================

def group_participants(data, column_mapping):
    """
    MAIN GROUPING ALGORITHM - Processes participants into optimized groups.

    PROCESSING ORDER (by priority):
    1. Filter out non-students (joiningAsStudent=False)
    2. Accountability Buddies (highest priority - explicit user requests)
    3. Solo Participants (user choice)
    4. Priority Same-Gender Grouping (females first, 5-member target, same location)
    5. Regular Algorithmic Grouping (remaining participants)
    6. Small Group Merging (optimization)

    Args:
        data: List of participant dictionaries
        column_mapping: Column name mappings

    Returns:
        tuple: (solo_groups, grouped, excluded_users, requested_groups, combined_group_info)
    """
    solo_groups = []
    grouped = defaultdict(list)
    group_counter = 1
    
    # Create dynamic email mapping
    email_mapping = create_email_mapping(data, column_mapping)
    
    # User tracking for diagnostics
    user_tracking = {}
    original_count = len(data)
    
    # Helper function to get value safely
    def get_value(row, key, default=''):
        if column_mapping and key in column_mapping:
            if isinstance(row, dict):
                return row.get(column_mapping[key], default)
            else:
                # For list format, we can't use column mapping
                return default
        else:
            # Fallback to old format (list indices)
            if isinstance(row, list):
                if key == 'go_solo':
                    return row[20] if len(row) > 20 else default
                elif key == 'user_id':
                    return row[0] if len(row) > 0 else default
                elif key == 'gender_preference':
                    return row[10] if len(row) > 10 else default
                elif key == 'gender_identity':
                    return row[3] if len(row) > 3 else default
                elif key == 'sex':
                    return row[7] if len(row) > 7 else default
                elif key == 'residing_ph':
                    return row[8] if len(row) > 8 else default
                elif key == 'country':
                    return row[16] if len(row) > 16 else default
                elif key == 'province':
                    return row[17] if len(row) > 17 else default
                elif key == 'city':
                    return row[18] if len(row) > 18 else default
                elif key == 'state':
                    return row[19] if len(row) > 19 else default
            return default
    
    # Initialize tracking for all users
    for i, row in enumerate(data):
        user_id = get_value(row, 'user_id', f'Row_{i}')
        # Convert user_id to string for consistent comparison
        user_id_str = str(user_id).strip() if user_id else f'Row_{i}'
        email = get_value(row, 'email', '')
        user_tracking[user_id_str] = {
            'email': email,
            'status': 'original',
            'reason': 'Initial data',
            'row_data': row
        }
    
    # Filter out participants where joiningAsStudent is False (but keep NaN/missing values)
    excluded_users = []  # Track excluded users to include them later
    if column_mapping and 'joining_as_student' in column_mapping:
        joining_col = column_mapping['joining_as_student']
        # Keep participants where joiningAsStudent is True or NaN/missing
        filtered_data = []
        excluded_count = 0
        for row in data:
            user_id = get_value(row, 'user_id', 'Unknown')
            joining_value = get_value(row, 'joining_as_student', 'True')
            # Convert to string and check if it's explicitly False
            joining_str = str(joining_value).strip().lower()
            if joining_str in ['false', '0', '0.0', 'no']:
                excluded_count += 1
                excluded_users.append(row)  # Add to excluded list
                user_id_str = str(user_id).strip() if user_id else 'Unknown'
                if user_id_str in user_tracking:
                    user_tracking[user_id_str]['status'] = 'excluded'
                    user_tracking[user_id_str]['reason'] = f'joiningAsStudent = {joining_value}'
            else:
                # Keep if True, NaN, or any other value (including missing)
                filtered_data.append(row)
        
        data = filtered_data
    
    # ============================================================================
    # PHASE 1: ACCOUNTABILITY BUDDIES (Highest Priority)
    # ============================================================================
    # Process users who explicitly requested to be grouped with specific buddies.
    # Uses graph-based clustering to find connected components of mutual buddies.
    requested_groups = []
    accountability_count = 0
    
    # First pass: collect all participants with non-empty accountabilityBuddies
    accountability_participants = []
    for row in data:
        accountability_buddies = get_value(row, 'accountability_buddies', '')
        has_accountability_buddies = get_value(row, 'has_accountability_buddies', '0')
        user_id = get_value(row, 'user_id', 'Unknown')
        
        # Check if has_accountability_buddies is True/1
        has_buddies = str(has_accountability_buddies).strip().lower() in ['1', '1.0', 'true', 'yes']
        
        # Check if accountability_buddies field has valid data
        has_buddy_data = False
        if accountability_buddies:
            accountability_str = str(accountability_buddies).strip()
            # Check if it's not empty and not just None values
            if accountability_str not in ['', 'None', 'nan', '[None]', '[None, None]', "{'1': None}"]:
                # For string representations of lists, check if there are actual email addresses
                emails = extract_emails_from_accountability_buddies(accountability_buddies, email_mapping)
                has_buddy_data = len(emails) > 0
        
        # Include users with has_buddies=True, even if they don't have buddy data (they might be referenced by others)
        if has_buddies:
            accountability_participants.append(row)
            user_id_str = str(user_id).strip() if user_id else 'Unknown'
            if user_id_str in user_tracking:
                user_tracking[user_id_str]['status'] = 'accountability_buddies'
                user_tracking[user_id_str]['reason'] = 'Has accountability buddies'
    
    # Second pass: collect users who are referenced as buddies by others
    referenced_buddies = set()
    for row in data:
        accountability_buddies = get_value(row, 'accountability_buddies', '')
        if accountability_buddies:
            emails = extract_emails_from_accountability_buddies(accountability_buddies, email_mapping)
            referenced_buddies.update(emails)
    
    # Add users who are referenced as buddies but not already in accountability_participants
    for row in data:
        user_email = normalize_email(get_value(row, 'email', ''), email_mapping)
        user_id = get_value(row, 'user_id', 'Unknown')
        
        if user_email in referenced_buddies:
            # Check if this user is already in accountability_participants
            already_included = any(
                normalize_email(get_value(acc_user, 'email', ''), email_mapping) == user_email 
                for acc_user in accountability_participants
            )
            
            if not already_included:
                accountability_participants.append(row)
                user_id_str = str(user_id).strip() if user_id else 'Unknown'
                if user_id_str in user_tracking:
                    user_tracking[user_id_str]['status'] = 'accountability_buddies'
                    user_tracking[user_id_str]['reason'] = 'Referenced as buddy by others'
    
    # Create a mapping of email to user data for quick lookup
    email_to_user = {}
    for row in data:
        # Use the column mapping to find email
        email = get_value(row, 'email', '')

        if email and '@' in email:
            normalized_email = normalize_email(email, email_mapping)
            email_to_user[normalized_email] = row
    
    # Pre-process: Group users with mutual buddies and their referenced buddies
    # Use a more comprehensive approach to find all connected groups
    mutual_buddy_groups = []
    processed_users = set()
    assigned_users = set()  # Track users already assigned to requested groups
    
    # Create a graph of all accountability buddy relationships
    buddy_graph = {}
    for participant in accountability_participants:
        participant_email = normalize_email(get_value(participant, 'email', ''), email_mapping)
        accountability_buddies = get_value(participant, 'accountability_buddies', '')
        requested_emails = extract_emails_from_accountability_buddies(accountability_buddies, email_mapping)
        
        if participant_email not in buddy_graph:
            buddy_graph[participant_email] = set()
        
        # Add direct references
        for email in requested_emails:
            if email in email_to_user:
                buddy_graph[participant_email].add(email)
        
        # Also add reverse references (users who reference this participant)
        for other_participant in accountability_participants:
            other_email = normalize_email(get_value(other_participant, 'email', ''), email_mapping)
            other_buddies = get_value(other_participant, 'accountability_buddies', '')
            other_requested_emails = extract_emails_from_accountability_buddies(other_buddies, email_mapping)
            
            if participant_email in other_requested_emails:
                if other_email not in buddy_graph:
                    buddy_graph[other_email] = set()
                buddy_graph[other_email].add(participant_email)
    
    # Find all connected components using DFS
    def find_connected_component(start_email, visited):
        """Find all emails connected to start_email through buddy relationships"""
        if start_email in visited:
            return set()
        
        visited.add(start_email)
        component = {start_email}
        
        if start_email in buddy_graph:
            for buddy_email in buddy_graph[start_email]:
                if buddy_email in email_to_user:  # Only include emails that exist in our data
                    component.update(find_connected_component(buddy_email, visited))
        
        # Also check for reverse connections (users who reference this email)
        for other_email, buddies in buddy_graph.items():
            if start_email in buddies and other_email not in visited:
                component.update(find_connected_component(other_email, visited))
        
        return component
    
    def split_large_component_by_direct_connections(connected_emails, max_group_size=7):
        """
        Split large accountability buddy groups into smaller groups.

        When connected components exceed max_group_size (7), split them evenly
        into smaller groups to ensure manageable group sizes.

        Args:
            connected_emails: Set of emails in the connected component
            max_group_size: Maximum allowed group size (default 7)

        Returns:
            List of email groups, each containing <= max_group_size emails
        """
        if len(connected_emails) <= max_group_size:
            return [list(connected_emails)]

        # Split evenly into groups of max_group_size
        final_groups = []
        emails_list = list(connected_emails)
        for i in range(0, len(emails_list), max_group_size):
            group = emails_list[i:i + max_group_size]
            final_groups.append(group)

        return final_groups
    
    # Alternative approach: ensure all referenced users are included
    def ensure_referenced_users_included():
        """Ensure that all users who are referenced by others are included in the same groups"""
        # Create a mapping of referenced users to their referrers
        referenced_to_referrers = {}
        for email, buddies in buddy_graph.items():
            for buddy_email in buddies:
                if buddy_email in email_to_user:
                    if buddy_email not in referenced_to_referrers:
                        referenced_to_referrers[buddy_email] = set()
                    referenced_to_referrers[buddy_email].add(email)
        
        # For each referenced user, ensure they're in the same group as their referrers
        for referenced_email, referrers in referenced_to_referrers.items():
            if referenced_email not in buddy_graph:  # User has no outgoing edges
                # Add this user to the buddy graph with connections to all referrers
                if referenced_email not in buddy_graph:
                    buddy_graph[referenced_email] = set()
                buddy_graph[referenced_email].update(referrers)
    
    # Call the function to ensure referenced users are included
    ensure_referenced_users_included()
    
    # Find all connected components
    visited = set()
    for participant in accountability_participants:
        participant_email = normalize_email(get_value(participant, 'email', ''), email_mapping)
        
        if participant_email not in visited:
            # Find all users connected to this participant
            connected_emails = find_connected_component(participant_email, visited)
            
            if len(connected_emails) > 1:
                # Split large components into smaller groups prioritizing direct connections
                if len(connected_emails) > 7:
                    # Use the splitting function for large components
                    email_groups = split_large_component_by_direct_connections(connected_emails, max_group_size=7)
                    
                    for email_group in email_groups:
                        if len(email_group) > 1:
                            mutual_group = []
                            for email in email_group:
                                if email in email_to_user:
                                    user = email_to_user[email]
                                    mutual_group.append(user)
                                    processed_users.add(email)
                            
                            if len(mutual_group) > 1:
                                mutual_buddy_groups.append(mutual_group)
                else:
                    # For smaller components, create a single group as before
                    mutual_group = []
                    for email in connected_emails:
                        if email in email_to_user:
                            user = email_to_user[email]
                            mutual_group.append(user)
                            processed_users.add(email)
                    
                    if len(mutual_group) > 1:
                        mutual_buddy_groups.append(mutual_group)
            else:
                # Single user - mark as processed but don't create a group yet
                processed_users.add(participant_email)
    
    # Process mutual buddy groups first
    for mutual_group in mutual_buddy_groups:
        if len(mutual_group) > 1:
            # Create a group for these mutual buddies
            requested_groups.append(mutual_group)
            accountability_count += len(mutual_group)
            
            # Mark all members as assigned
            for member in mutual_group:
                member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                assigned_users.add(member_email)
    
    # Process remaining accountability participants (those not in mutual groups)
    remaining_participants = []
    for participant in accountability_participants:
        participant_email = normalize_email(get_value(participant, 'email', ''), email_mapping)
        if participant_email not in assigned_users:
            remaining_participants.append(participant)
    
    # Process remaining participants with the original logic
    processed_requests = set()  # Track processed requests to avoid duplicates
    
    for participant in remaining_participants:
        accountability_buddies = get_value(participant, 'accountability_buddies', '')
        user_id = get_value(participant, 'user_id', 'Unknown')
        participant_email = normalize_email(get_value(participant, 'email', ''), email_mapping)
        
        # Clean and extract emails from accountabilityBuddies
        requested_emails = extract_emails_from_accountability_buddies(accountability_buddies, email_mapping)
        
        if requested_emails:
            # Create a unique key for this request to avoid duplicates
            request_key = ','.join(sorted(requested_emails))
            
            if request_key not in processed_requests:
                processed_requests.add(request_key)
                
                # Check if any requested buddies are already in existing groups
                buddies_in_existing_groups = []
                available_buddies = []
                existing_group_with_buddies = None
                
                for email in requested_emails:
                    if email in email_to_user:
                        buddy_user = email_to_user[email]
                        buddy_email = normalize_email(get_value(buddy_user, 'email', ''), email_mapping)
                        
                        # Check if this buddy is already assigned to a requested group
                        if buddy_email in assigned_users:
                            buddies_in_existing_groups.append(email)
                            
                            # Find which existing group contains this buddy
                            for i, existing_group in enumerate(requested_groups):
                                existing_emails = [normalize_email(get_value(member, 'email', ''), email_mapping) for member in existing_group]
                                if buddy_email in existing_emails:
                                    existing_group_with_buddies = i
                                    break
                        else:
                            available_buddies.append(email)
                    else:
                        pass
                
                # If buddies are in existing groups, add user to that group
                if buddies_in_existing_groups and existing_group_with_buddies is not None:
                    existing_group = requested_groups[existing_group_with_buddies]
                    
                    # Check if the group has space (max 5 members)
                    if len(existing_group) < 5:
                        # Check if participant is already in the group to prevent duplicates
                        participant_email_normalized = normalize_email(get_value(participant, 'email', ''), email_mapping)
                        existing_emails = [normalize_email(get_value(member, 'email', ''), email_mapping) for member in existing_group]
                        if participant_email_normalized not in existing_emails:
                            # Add participant to existing group
                            existing_group.append(participant)
                            assigned_users.add(participant_email)
                            accountability_count += 1
                            
                            # Also add any available buddies to the same group if there's space
                            for email in available_buddies:
                                if email in email_to_user:
                                    buddy_user = email_to_user[email]
                                    buddy_email = normalize_email(get_value(buddy_user, 'email', ''), email_mapping)
                                    if buddy_email not in assigned_users and len(existing_group) < 5:
                                        existing_group.append(buddy_user)
                                        assigned_users.add(buddy_email)
                                        accountability_count += 1
                        
                    else:
                        # Create a new group for this user since existing group is full
                        group_members = [participant]
                        assigned_users.add(participant_email)
                        
                        # Add any available buddies
                        for email in available_buddies:
                            if email in email_to_user:
                                buddy_user = email_to_user[email]
                                buddy_email = normalize_email(get_value(buddy_user, 'email', ''), email_mapping)
                                if buddy_email not in assigned_users:
                                    group_members.append(buddy_user)
                                    assigned_users.add(buddy_email)
                        
                        if group_members:
                            requested_groups.append(group_members)
                            accountability_count += len(group_members)
                            
                
                # Create a new group with available buddies (if any)
                elif available_buddies:
                    # Build the group: requester + all available buddies
                    group_members = [participant]  # Start with the requester
                    assigned_users.add(participant_email)  # Mark requester as assigned
                    
                    found_buddies = []
                    missing_buddies = []
                    
                    for email in available_buddies:
                        buddy_user = email_to_user[email]
                        buddy_email = normalize_email(get_value(buddy_user, 'email', ''), email_mapping)
                        
                        # Check if buddy is already in the group to prevent duplicates
                        if buddy_email not in [normalize_email(get_value(member, 'email', ''), email_mapping) for member in group_members]:
                            group_members.append(buddy_user)
                            assigned_users.add(buddy_email)  # Mark buddy as assigned
                            found_buddies.append(email)
                    
                    # Add missing buddies to the list
                    for email in requested_emails:
                        if email not in email_to_user:
                            missing_buddies.append(email)
                    
                    if group_members:
                        requested_groups.append(group_members)
                        accountability_count += len(group_members)
                        
                
                # Create a group for the requester even if no buddies are available
                else:
                    # Create a group with just the requester (missing buddies)
                    group_members = [participant]
                    assigned_users.add(participant_email)
                    
                    if group_members:
                        requested_groups.append(group_members)
                        accountability_count += len(group_members)
    
    # Final pass: ensure all remaining accountability participants are assigned to groups
    # This catches any users that might have been missed in the previous processing
    for participant in remaining_participants:
        participant_email = normalize_email(get_value(participant, 'email', ''), email_mapping)
        
        # Skip if already assigned
        if participant_email in assigned_users:
            continue

        # Create a solo group for this user
        group_members = [participant]
        assigned_users.add(participant_email)
        requested_groups.append(group_members)
        accountability_count += 1

    
    # Move single-member requested groups to regular groups BEFORE regular grouping
    single_member_requested_groups = []
    multi_member_requested_groups = []
    
    for group in requested_groups:
        if len(group) == 1:
            # Move ALL single-member groups to regular groups
            # This ensures no users are left isolated in single-member groups
            user = group[0]
            # Single member group - move to regular groups
            single_member_requested_groups.append(user)

            # Update user tracking
            user_email = normalize_email(get_value(user, 'email', ''), email_mapping)
            for user_id, info in user_tracking.items():
                if info.get('email') == user_email:
                    info['status'] = 'regular_grouping'
                    info['reason'] = 'Moved from single-member requested group to regular grouping'
                    break

            # Remove from assigned_users so they can go through regular grouping
            assigned_users.discard(user_email)
        else:
            # Multi-member group - keep as requested group
            multi_member_requested_groups.append(group)
    
    # Keep original requested groups without combining
    # No combining logic - keep groups as they were originally formed
    
    # ============================================================================
    # PHASE 2: SOLO PARTICIPANTS (User Choice)
    # ============================================================================
    # Users who explicitly chose to work individually (go_solo=True)
    # These participants are placed in single-member groups
    solo_count = 0
    # Remove accountability participants and already assigned users from data for solo processing
    remaining_data = []
    for row in data:
        user_email = normalize_email(get_value(row, 'email', ''), email_mapping)
        # Skip if user is already assigned to requested groups
        if user_email not in assigned_users:
            remaining_data.append(row)
    
    # Add single-member requested group users to remaining_data for processing
    for user in single_member_requested_groups:
        user_email = normalize_email(get_value(user, 'email', ''), email_mapping)
        # Only add if not already in remaining_data
        if not any(normalize_email(get_value(row, 'email', ''), email_mapping) == user_email for row in remaining_data):
            remaining_data.append(user)
    
    for row in remaining_data:
        go_solo_value = str(get_value(row, 'go_solo', '0')).strip()
        user_id = get_value(row, 'user_id', 'Unknown')
        user_email = normalize_email(get_value(row, 'email', ''), email_mapping)
        # Handle various formats: '1', '1.0', 'True', 'true'
        if go_solo_value.lower() in ['1', '1.0', 'true']:
            solo_groups.append([row])
            solo_count += 1
            assigned_users.add(user_email)  # Mark as assigned
            user_id_str = str(user_id).strip() if user_id else 'Unknown'
            if user_id_str in user_tracking:
                user_tracking[user_id_str]['status'] = 'solo'
                user_tracking[user_id_str]['reason'] = 'go_solo = True'
    
    # ============================================================================
    # PHASE 3: REGULAR GROUPING (Algorithmic - Gender + Geography)
    # ============================================================================
    # Process remaining participants using hierarchical grouping algorithm:
    # 1. Separate by gender preference (same_gender vs no_preference)
    # 2. Within each gender group, separate by geography (PH vs International)
    # 3. Philippines: Province → City → Same-city groups
    # 4. International: Country → State → Timezone regions
    # 5. Optimize group sizes (3-5 members per group)

    non_solo = [row for row in remaining_data if str(get_value(row, 'go_solo', '0')).strip().lower() not in ['1', '1.0', 'true']]

    # ============================================================================
    # STEP 1: PRIORITY GROUPING - Same Gender First, Females First
    # ============================================================================
    # Process same_gender participants first, females before males
    # Target group size: 5 members from same location
    # Fill gaps with no_preference participants from same location

    # Separate participants by sex and preference
    sex_preference_groups = {
        'female': {'same_gender': [], 'no_preference': []},
        'male': {'same_gender': [], 'no_preference': []}
    }

    for row in non_solo:
        sex = str(get_value(row, 'sex', '')).lower()
        gender_pref = str(get_value(row, 'gender_preference', '')).lower()

        # Only process female/male (ignore other genders for now)
        if sex in ['female', 'male']:
            if gender_pref == 'same_gender':
                sex_preference_groups[sex]['same_gender'].append(row)
            elif gender_pref == 'no_preference':
                sex_preference_groups[sex]['no_preference'].append(row)
            # Ignore 'other' preferences for now - they'll be handled later

    # Process females first, then males
    for current_sex in ['female', 'male']:
        same_gender_participants = sex_preference_groups[current_sex]['same_gender']
        no_preference_participants = sex_preference_groups[current_sex]['no_preference']

        # Group same_gender participants by location first
        location_groups = defaultdict(list)

        # Separate by Philippines vs International
        for participant in same_gender_participants:
            ph_val = str(get_value(participant, 'residing_ph', '0')).strip().lower()
            if ph_val in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
                # Philippines: use province_city combination
                province = str(get_value(participant, 'province', '')).strip()
                city = str(get_value(participant, 'city', '')).strip()
                location_key = f"PH_{province}_{city}"
            else:
                # International: use country_state combination
                country = str(get_value(participant, 'country', '')).strip()
                state = str(get_value(participant, 'state', '')).strip()
                location_key = f"INT_{country}_{state}"

            location_groups[location_key].append(participant)

        # Process each location group - create groups of 5 same_gender participants
        for location_key, participants in location_groups.items():
            # Sort participants for consistent ordering
            participants.sort(key=lambda p: get_value(p, 'user_id', ''))

            # Create groups of 5 from same_gender participants
            i = 0
            while i < len(participants):
                # Take up to 5 participants for this group
                group_size = min(5, len(participants) - i)
                group_members = participants[i:i + group_size]

                # If we have less than 5, try to fill with no_preference participants from same location
                if len(group_members) < 5:
                    # Find no_preference participants from same location
                    available_fillers = []
                    for filler in no_preference_participants:
                        ph_val = str(get_value(filler, 'residing_ph', '0')).strip().lower()
                        if ph_val in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
                            filler_province = str(get_value(filler, 'province', '')).strip()
                            filler_city = str(get_value(filler, 'city', '')).strip()
                            filler_location = f"PH_{filler_province}_{filler_city}"
                        else:
                            filler_country = str(get_value(filler, 'country', '')).strip()
                            filler_state = str(get_value(filler, 'state', '')).strip()
                            filler_location = f"INT_{filler_country}_{filler_state}"

                        if filler_location == location_key:
                            available_fillers.append(filler)

                    # Add fillers to reach target of 5 (but don't exceed)
                    fillers_needed = 5 - len(group_members)
                    fillers_to_add = available_fillers[:fillers_needed]

                    # Remove used fillers from available pool
                    for filler in fillers_to_add:
                        no_preference_participants.remove(filler)

                    group_members.extend(fillers_to_add)

                # Create the group name based on location
                if location_key.startswith('PH_'):
                    parts = location_key.split('_', 2)
                    if len(parts) >= 3:
                        province, city = parts[1], parts[2]
                        location_info = f"Province: {province}, City: {city}"
                    else:
                        location_info = "Philippines"
                else:  # INT_
                    parts = location_key.split('_', 2)
                    if len(parts) >= 3:
                        country, state = parts[1], parts[2]
                        location_info = f"Country: {country}, State: {state}"
                    else:
                        location_info = "International"

                group_name = f"Group {group_counter} ({current_sex}, same_gender, {location_info})"
                grouped[group_name] = group_members

                # Mark all members as assigned
                for member in group_members:
                    member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                    assigned_users.add(member_email)

                group_counter += 1
                i += len(group_members)  # Move past the participants we used

        # Handle remaining same_gender participants that couldn't form groups of 3+
        remaining_same_gender = []
        for location_key, participants in location_groups.items():
            remaining_same_gender.extend(participants)

        # If we have remaining participants, create smaller groups or add to mixed groups later
        # For now, they'll be handled in the regular algorithmic grouping below

    # After processing priority same_gender groups, handle remaining participants
    # with regular algorithmic grouping
    remaining_participants = []
    for row in non_solo:
        user_email = normalize_email(get_value(row, 'email', ''), email_mapping)
        if user_email not in assigned_users:
            remaining_participants.append(row)

    # ============================================================================
    # STEP 3: REGULAR ALGORITHMIC GROUPING (for remaining participants)
    # ============================================================================
    # Apply standard geographic grouping to participants not assigned to priority groups

    # Step 4: Group remaining participants by gender preference
    gender_pref_groups = defaultdict(list)

    for row in remaining_participants:
        gender_pref = str(get_value(row, 'gender_preference', '')).lower()
        user_id = get_value(row, 'user_id', 'Unknown')

        # Determine grouping key based on gender preferences
        if gender_pref == 'same_gender':
            # STRICT GENDER SEPARATION: Same biological sex only
            sex = str(get_value(row, 'sex', '')).lower()
            gender_identity = str(get_value(row, 'gender_identity', '')).upper()

            if gender_identity == 'LGBTQ+':
                # LGBTQ+ participants grouped by biological sex for same-gender preference
                gender_key = f"lgbtq+_{sex}"
            else:
                # Regular participants: group by biological sex
                gender_key = sex
        elif gender_pref == 'no_preference':
            # MIXED GENDER: Allow any gender combination
            gender_key = 'no_preference'
        else:
            # Unknown preference: separate group
            gender_key = 'other'

        gender_pref_groups[gender_key].append(row)

    # Step 5: Within each gender group, apply geographic grouping
    for gender_key, rows in gender_pref_groups.items():
        # Separate Philippines vs International residents
        ph_rows = []
        non_ph_rows = []
        
        for r in rows:
            ph_val = str(get_value(r, 'residing_ph', '0')).strip().lower()
            if ph_val in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
                ph_rows.append(r)
            elif ph_val in ['0', '0.0', 'false', 'no']:
                non_ph_rows.append(r)
            else:
                # For unknown values, treat as international
                non_ph_rows.append(r)
        
    # PHILIPPINES GROUPING: Province → City → Same-city groups
    # Prioritizes keeping participants from same city together (Step 5)
        province_groups = defaultdict(list)
        for r in ph_rows:
            province = get_value(r, 'province', 'Unknown Province')
            # Normalize province name
            province_norm = province.strip().lower() if isinstance(province, str) else str(province).strip().lower()
            province_groups[province_norm].append(r)
        
        # Sort provinces by Philippines regions (Luzon, Visayas, Mindanao)
        sorted_provinces = []
        for province_norm, province_members in province_groups.items():
            # Get the original province name from the first member for sorting
            original_province = get_value(province_members[0], 'province', 'Unknown Province')
            region = get_philippines_region(original_province)
            sorted_provinces.append((original_province, province_norm, province_members, region))
        
        # Sort by region first (Luzon, Visayas, Mindanao), then by province name within each region
        region_order = {'luzon': 1, 'visayas': 2, 'mindanao': 3, 'unknown': 4}
        sorted_provinces.sort(key=lambda x: (region_order.get(x[3], 5), str(x[0]).lower() if x[0] else ''))
        
        for original_province, province_norm, province_members, region in sorted_provinces:
            # Use the original province name for display
            province = original_province
            # Further group by city within each province
            city_groups = defaultdict(list)
            for r in province_members:
                city = get_value(r, 'city', 'Unknown City')
                # Normalize city name
                city_norm = city.strip().lower() if isinstance(city, str) else str(city).strip().lower()
                city_groups[city_norm].append(r)
            
            # --- SORT CITIES ALPHABETICALLY ---
            sorted_city_names = sorted(city_groups.keys())
            
            # --- NEW LOGIC: Prioritize same-city groups from entire province pool ---
            # Collect all participants from this province
            all_province_members = []
            for city_norm in sorted_city_names:  # Use sorted city names
                members = city_groups[city_norm]
                all_province_members.extend(members)
            
            # Group by city within the province
            city_members = defaultdict(list)
            for member in all_province_members:
                city = get_value(member, 'city', 'Unknown City')
                city_norm = city.strip().lower() if isinstance(city, str) else str(city).strip().lower()
                city_members[city_norm].append(member)
            
            # First, create complete groups (5 members) from each city
            remaining_by_city = {}
            for city_norm, members in city_members.items():
                # Create complete groups of 5 from this city
                i = 0
                while i + 5 <= len(members):
                    group_members = members[i:i+5]
                    location_info = f"Province: {province}, City: {city_norm}"
                    grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group_members
                    # Mark all members as assigned
                    for member in group_members:
                        member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                        assigned_users.add(member_email)
                    group_counter += 1
                    i += 5
                
                # Keep remaining members from this city
                if i < len(members):
                    remaining_by_city[city_norm] = members[i:]
            
            # Now handle remaining members - prioritize same-city groups
            if remaining_by_city:
                # First, try to form same-city groups from remaining members
                for city_norm, members in list(remaining_by_city.items()):
                    if len(members) >= 5:
                        # Can form a complete group from this city
                        group_members = members[:5]
                        location_info = f"Province: {province}, City: {city_norm}"
                        grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group_members
                        # Mark all members as assigned
                        for member in group_members:
                            member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                            assigned_users.add(member_email)
                        group_counter += 1
                        remaining_by_city[city_norm] = members[5:]
                    elif len(members) == 0:
                        del remaining_by_city[city_norm]
                
                # Collect all final remaining members (less than 5 per city)
                final_remaining = []
                for members in remaining_by_city.values():
                    final_remaining.extend(members)
                
                # Create mixed-city groups from final remaining - keep city-units together
                if final_remaining:
                    # Group final remaining by city - use remaining_by_city directly
                    final_by_city = []
                    for city, members in remaining_by_city.items():
                        if members:  # Only add non-empty city units
                            final_by_city.append(members)
                    
                    # Greedily combine city-units into groups of up to 5, never splitting a city-unit
                    i = 0
                    while i < len(final_by_city):
                        group = []
                        while i < len(final_by_city) and len(group) + len(final_by_city[i]) <= 5:
                            group.extend(final_by_city[i])
                            i += 1
                        if group:
                            # Check if all from same city
                            cities_in_group = set()
                            for member in group:
                                city = get_value(member, 'city', 'Unknown City')
                                cities_in_group.add(city)
                            
                            if len(cities_in_group) == 1:
                                # All from same city
                                city_name = list(cities_in_group)[0]
                                location_info = f"Province: {province}, City: {city_name}"
                            else:
                                # Mixed cities
                                location_info = f"Province: {province}, Mixed Cities"
                            
                            grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group
                            # Mark all members as assigned
                            for member in group:
                                member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                                assigned_users.add(member_email)
                            group_counter += 1
        
    # INTERNATIONAL GROUPING: Country → State → Timezone regions
    # Uses timezone clustering for geographic proximity (Step 5)
        if non_ph_rows:
            # Group by country first, then state, then timezone region
            country_groups = defaultdict(list)
            for r in non_ph_rows:
                country = get_value(r, 'country', 'Unknown Country')
                # Normalize country name
                country_norm = country.strip().lower() if isinstance(country, str) else str(country).strip().lower()
                country_groups[country_norm].append(r)
            
            # Sort countries alphabetically
            sorted_countries = sorted(country_groups.items())
            
            for country_norm, country_members in sorted_countries:
                # Use the original country name for display
                country = get_value(country_members[0], 'country', 'Unknown Country')
                # Further group by state within each country
                state_groups = defaultdict(list)
                for r in country_members:
                    state = get_value(r, 'state', 'Unknown State')
                    # Normalize state name
                    state_norm = state.strip().lower() if isinstance(state, str) else str(state).strip().lower()
                    state_groups[state_norm].append(r)
                
                # Sort states alphabetically
                sorted_state_names = sorted(state_groups.keys())
                
                # Create groups from each state
                for state_norm in sorted_state_names:
                    members = state_groups[state_norm]
                    # Use the original state name for display
                    state = get_value(members[0], 'state', 'Unknown State')
                    
                    # Create complete groups of 5 from this state
                    i = 0
                    while i + 5 <= len(members):
                        group_members = members[i:i+5]
                        location_info = f"Country: {country}, State: {state}"
                        grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group_members
                        # Mark all members as assigned
                        for member in group_members:
                            member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                            assigned_users.add(member_email)
                        group_counter += 1
                        i += 5
                    
                    # Handle remaining members from this state
                    if i < len(members):
                        remaining_members = members[i:]
                        if len(remaining_members) >= 5:
                            # Can form a complete group
                            group_members = remaining_members[:5]
                            location_info = f"Country: {country}, State: {state}"
                            grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group_members
                            # Mark all members as assigned
                            for member in group_members:
                                member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                                assigned_users.add(member_email)
                            group_counter += 1
                            remaining_members = remaining_members[5:]
                        
                        # Add remaining members to a mixed group if any
                        if remaining_members:
                            # Check if we can combine with other remaining members from other states/countries
                            all_remaining_international = []
                            for other_country_norm, other_country_members in country_groups.items():
                                if other_country_norm != country_norm:
                                    for other_state_norm, other_state_members in state_groups.items():
                                        if other_state_norm != state_norm:
                                            # Get remaining members from other states
                                            other_remaining = []
                                            for other_member in other_state_members:
                                                other_email = normalize_email(get_value(other_member, 'email', ''), email_mapping)
                                                if other_email not in assigned_users:
                                                    other_remaining.append(other_member)
                                            all_remaining_international.extend(other_remaining)
                            
                            # Combine with other remaining international members
                            combined_remaining = remaining_members + all_remaining_international
                            
                            if combined_remaining:
                                # Create groups from combined remaining
                                i = 0
                                while i + 5 <= len(combined_remaining):
                                    group_members = combined_remaining[i:i+5]
                                    location_info = f"International Mixed"
                                    grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = group_members
                                    # Mark all members as assigned
                                    for member in group_members:
                                        member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                                        assigned_users.add(member_email)
                                    group_counter += 1
                                    i += 5
                                
                                # Handle final remaining (less than 5)
                                if i < len(combined_remaining):
                                    final_group = combined_remaining[i:]
                                    location_info = f"International Mixed"
                                    grouped[f"Group {group_counter} ({gender_key}, {location_info})"] = final_group
                                    # Mark all members as assigned
                                    for member in final_group:
                                        member_email = normalize_email(get_value(member, 'email', ''), email_mapping)
                                        assigned_users.add(member_email)
                                    group_counter += 1
    
    # Merge small groups based on geographic proximity
    grouped = merge_small_groups(grouped, column_mapping)
    
    # Generate diagnostic report
    generate_diagnostic_report(user_tracking, original_count, solo_groups, grouped, excluded_users, multi_member_requested_groups, column_mapping)
    
    # Generate missing users analysis
    generate_missing_users_analysis(user_tracking, original_count, solo_groups, grouped, excluded_users, multi_member_requested_groups, column_mapping)
    
    # Check for duplicates
    check_for_duplicates(solo_groups, grouped, excluded_users, multi_member_requested_groups, column_mapping)
    
    # Return the updated groups - use original requested groups without combining
    return solo_groups, grouped, excluded_users, multi_member_requested_groups, {}

def save_to_excel(solo_groups, grouped, filename_or_buffer, column_mapping, excluded_users=None, requested_groups=None, combined_group_info=None):
    """
    Save all groups to formatted Excel file with color coding and structure.

    OUTPUT STRUCTURE (by sheet row order):
    1. Requested Groups (Accountability Buddies) - Green highlight if ≥5 members
    2. Solo Groups - Individual participants
    3. Regular Groups - Light blue highlight if ≥5 members + same location
    4. Excluded Users - Participants who opted out (joiningAsStudent=False)

    EXCEL COLUMNS:
    - Group Name, User ID 1-7, Name 1-7, Location 1-7, Coach 1-7
    - Gender Identity, Sex, Residing in PH, Gender Preference
    - Country, Province, City, State, Previous Coach Name

    VISUAL FORMATTING:
    - User IDs: Sex-based fill colors (blue=male, pink=female)
    - Special colors: Green fill for 'get_bigger' goal, maroon font for LGBTQ+
    - Bold text: Same-gender preference groups
    - Underlined text: Users with accountability buddies
    - Group highlighting: Green (requested ≥5), Blue (regular ≥5 + same location)

    Args:
        solo_groups: List of single-member groups
        grouped: Dict of {group_name: [members]} for regular groups
        filename_or_buffer: Output file path or BytesIO buffer
        column_mapping: Column name mappings
        excluded_users: Users who opted out
        requested_groups: Accountability buddy groups
        combined_group_info: Information about combined groups
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Members"
    ws.append([
        "Group Name",
        "User ID 1", "Name 1", "Location 1", "Coach and Age 1",
        "User ID 2", "Name 2", "Location 2", "Coach and Age 2",
        "User ID 3", "Name 3", "Location 3", "Coach and Age 3",
        "User ID 4", "Name 4", "Location 4", "Coach and Age 4",
        "User ID 5", "Name 5", "Location 5", "Coach and Age 5",
        "User ID 6", "Name 6", "Location 6", "Coach and Age 6",
        "User ID 7", "Name 7", "Location 7", "Coach and Age 7",
        "Gender Identity", "Sex", "Residing in PH", "Gender Preference", "Country", "Province", "City", "State",
        "Previous Coach Name"
    ])
    
    # Write requested groups (accountability buddies)
    if requested_groups:
        # Sort requested groups by descending order of number of users
        sorted_requested_groups = sorted(requested_groups, key=lambda g: len(g), reverse=True)
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        group_row_indices = []
        for idx, group in enumerate(sorted_requested_groups, 1):
            # --- SORT small group members ---
            if len(group) < 7:
                group = sorted(group, key=lambda m: (
                    m.get(column_mapping.get('user_id'), ''),
                    m.get(column_mapping.get('name'), ''),
                    m.get(column_mapping.get('city'), ''),
                    m.get(column_mapping.get('previous_coach_name'), '')
                ))
            
            # All requested groups are accountability buddy groups
            first_member = group[0]
            accountability_buddies = first_member.get(column_mapping.get('accountability_buddies'), '')

            # Check if this is a combined group
            is_combined_group = False
            combined_info = ""
            if combined_group_info:
                # Get emails of current group members for comparison
                current_group_emails = set()
                for member in group:
                    email = member.get(column_mapping.get('email'), '')
                    if email and '@' in email:
                        current_group_emails.add(email.lower().strip())
                
                for group_info in combined_group_info.values():
                    # Get emails of the combined group members
                    combined_group_emails = set()
                    for member in group_info['members']:
                        email = member.get(column_mapping.get('email'), '')
                        if email and '@' in email:
                            combined_group_emails.add(email.lower().strip())
                    
                    # Compare email sets
                    if current_group_emails == combined_group_emails:
                        is_combined_group = group_info['is_combined']
                        combined_info = group_info['combined_info']
                        break
            
            # Determine if this is a team group or accountability buddy group
            team_names = set()
            has_accountability = False
            for member in group:
                # Check team (assuming kaizen_client_type is the team)
                team = safe_get_value(member, column_mapping.get('kaizen_client_type', ''), '')
                if team:
                    team_names.add(str(team).strip())
                
                # Check accountability
                acc = safe_get_value(member, column_mapping.get('accountability_buddies', ''), '')
                if acc and str(acc).strip() not in ['', 'None', 'nan', 'NaN']:
                    has_accountability = True
            
            all_same_team = len(team_names) == 1
            all_no_accountability = not has_accountability
            team_name = list(team_names)[0] if all_same_team else ""
            
            if all_same_team and all_no_accountability and team_name:
                # This is a team group
                group_name = f"Team Group {idx} - {team_name} ({len(group)} members)"
                if is_combined_group and combined_info:
                    group_name += f" - {combined_info}"
                row = [group_name]
            else:
                # This is an accountability buddy group
                # Check for missing buddies
                missing_buddies_info = ""
                if accountability_buddies and str(accountability_buddies).strip() not in ['', 'None', 'nan']:
                    # Extract emails from accountability_buddies using the helper function
                    # Create a temporary email mapping for this function
                    temp_email_mapping = create_email_mapping([], {})
                    requested_emails = extract_emails_from_accountability_buddies(accountability_buddies, temp_email_mapping)
                    
                    # Get emails of current group members
                    group_emails = []
                    for member in group:
                        email = member.get(column_mapping.get('email'), '')
                        if email and '@' in email:
                            group_emails.append(email.lower().strip())
                    
                    # Find missing buddies
                    missing_buddies = [email for email in requested_emails if email not in group_emails]
                    
                    if missing_buddies:
                        missing_buddies_info = f" - Missing: {', '.join(missing_buddies[:3])}"
                        if len(missing_buddies) > 3:
                            missing_buddies_info += f" (+{len(missing_buddies)-3} more)"
                
                group_name = f"Requested Group {idx} ({len(group)} members){missing_buddies_info}"
                if is_combined_group and combined_info:
                    group_name += f" - {combined_info}"
                row = [group_name]
            
            # Add user data for each member (up to 7)
            for i in range(7):
                if i < len(group):
                    member = group[i]
                    location_display = format_location_display(member, column_mapping)
                    coach_name = safe_get_value(member, column_mapping.get('previous_coach_name', ''), '')
                    age_group = safe_get_value(member, column_mapping.get('age_group', ''), '')
                    # Format coach name with age group in parentheses
                    coach_with_age = coach_name
                    if coach_name and age_group:
                        coach_with_age = f"{coach_name} ({age_group})"
                    elif coach_name:
                        coach_with_age = coach_name
                    elif age_group:
                        coach_with_age = f"({age_group})"

                    row.extend([
                        member.get(column_mapping.get('user_id'), ''),
                        member.get(column_mapping.get('name'), ''),
                        location_display,
                        coach_with_age
                    ])
                else:
                    row.extend(["", "", "", ""])
            
            # Add extra info for the first member
            member = group[0]

            # Collect all coach names with age groups from group members
            coach_names = []
            for group_member in group:
                coach_name = safe_get_value(group_member, column_mapping.get('previous_coach_name', ''), '')
                age_group = safe_get_value(group_member, column_mapping.get('age_group', ''), '')
                if coach_name and str(coach_name).strip() not in ['', 'None', 'nan']:
                    # Format coach name with age group
                    coach_with_age = coach_name
                    if coach_name and age_group:
                        coach_with_age = f"{coach_name} ({age_group})"
                    elif coach_name:
                        coach_with_age = coach_name
                    elif age_group:
                        coach_with_age = f"({age_group})"
                    coach_names.append(str(coach_with_age).strip())

            # Combine coach names with "/" separator if they're different
            combined_coach_names = ' / '.join(sorted(set(coach_names))) if coach_names else ''
            
            row.extend([
                member.get(column_mapping.get('gender_identity'), ''),
                member.get(column_mapping.get('sex'), ''),
                member.get(column_mapping.get('residing_ph'), ''),
                member.get(column_mapping.get('gender_preference'), ''),
                member.get(column_mapping.get('country'), ''),
                member.get(column_mapping.get('province'), ''),
                member.get(column_mapping.get('city'), ''),
                member.get(column_mapping.get('state'), ''),
                combined_coach_names
            ])
            
            ws.append(row)
            group_row_indices.append((ws.max_row, len(group)))
            
            # Apply formatting
            for i in range(7):
                if i < len(group):
                    member = group[i]
                    gender_pref = member.get(column_mapping.get('gender_preference'), '')
                    kaizen_client_type = member.get(column_mapping.get('kaizen_client_type'), '')
                    sex = member.get(column_mapping.get('sex'), '')
                    gender_identity = member.get(column_mapping.get('gender_identity'), '')
                    has_accountability_buddies = member.get(column_mapping.get('has_accountability_buddies'), '')
                    current_goal = member.get(column_mapping.get('current_goal'), '')
                    apply_color_to_cell(ws.cell(row=ws.max_row, column=2 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=True)  # User ID
                    apply_color_to_cell(ws.cell(row=ws.max_row, column=3 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=False)  # Name
        
        # After all requested groups are written, apply green highlight to group name cell if group has 5 or more members
        for row_idx, group_size in group_row_indices:
            if group_size >= 5:
                ws.cell(row=row_idx, column=1).fill = green_fill
    
    # Write solo groups
    for idx, group in enumerate(solo_groups, 1):
        # --- SORT small group members ---
        if len(group) < 7:
            group = sorted(group, key=lambda m: (
                m.get(column_mapping.get('user_id'), ''),
                m.get(column_mapping.get('name'), ''),
                m.get(column_mapping.get('city'), ''),
                m.get(column_mapping.get('previous_coach_name'), '')
            ))
        row = [f"Solo {idx}"]
        for i in range(7):
            if i < len(group):
                member = group[i]
                location_display = format_location_display(member, column_mapping)
                coach_name = safe_get_value(member, column_mapping.get('previous_coach_name', ''), '')
                age_group = safe_get_value(member, column_mapping.get('age_group', ''), '')
                # Format coach name with age group in parentheses
                coach_with_age = coach_name
                if coach_name and age_group:
                    coach_with_age = f"{coach_name} ({age_group})"
                elif coach_name:
                    coach_with_age = coach_name
                elif age_group:
                    coach_with_age = f"({age_group})"

                row.extend([
                    member.get(column_mapping.get('user_id'), ''),
                    member.get(column_mapping.get('name'), ''),
                    location_display,
                    coach_with_age
                ])
            else:
                row.extend(["", "", "", ""])
        # Add extra info for the first member
        member = group[0]
        row.extend([
            member.get(column_mapping.get('gender_identity'), ''),
            member.get(column_mapping.get('sex'), ''),
            member.get(column_mapping.get('residing_ph'), ''),
            member.get(column_mapping.get('gender_preference'), ''),
            member.get(column_mapping.get('country'), ''),
            member.get(column_mapping.get('province'), ''),
            member.get(column_mapping.get('city'), ''),
            member.get(column_mapping.get('state'), ''),
            member.get(column_mapping.get('previous_coach_name'), '')
        ])
        ws.append(row)
        # Color code user_id and name cells for each member
        for i in range(7):
            if i < len(group):
                member = group[i]
                gender_pref = safe_get_value(member, column_mapping.get('gender_preference', ''), '')
                kaizen_client_type = safe_get_value(member, column_mapping.get('kaizen_client_type', ''), '')
                sex = safe_get_value(member, column_mapping.get('sex', ''), '')
                gender_identity = safe_get_value(member, column_mapping.get('gender_identity', ''), '')
                has_accountability_buddies = safe_get_value(member, column_mapping.get('has_accountability_buddies', ''), '')
                current_goal = safe_get_value(member, column_mapping.get('current_goal', ''), '')
                apply_color_to_cell(ws.cell(row=ws.max_row, column=2 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=True)  # User ID
                apply_color_to_cell(ws.cell(row=ws.max_row, column=3 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=False)  # Name
    
    # Write grouped participants
    # Track regular groups with 5 or more members for highlighting
    regular_group_row_indices = []
    for group_name, members in grouped.items():
        # --- SORT small group members ---
        if len(members) < 7:
            members = sorted(members, key=lambda m: (
                m.get(column_mapping.get('user_id'), ''),
                m.get(column_mapping.get('name'), ''),
                m.get(column_mapping.get('city'), ''),
                m.get(column_mapping.get('previous_coach_name'), '')
            ))
        row = [group_name]
        for i in range(7):
            if i < len(members):
                member = members[i]
                location_display = format_location_display(member, column_mapping)
                coach_name = safe_get_value(member, column_mapping.get('previous_coach_name', ''), '')
                age_group = safe_get_value(member, column_mapping.get('age_group', ''), '')
                # Format coach name with age group in parentheses
                coach_with_age = coach_name
                if coach_name and age_group:
                    coach_with_age = f"{coach_name} ({age_group})"
                elif coach_name:
                    coach_with_age = coach_name
                elif age_group:
                    coach_with_age = f"({age_group})"

                row.extend([
                    member.get(column_mapping.get('user_id'), ''),
                    member.get(column_mapping.get('name'), ''),
                    location_display,
                    coach_with_age
                ])
            else:
                row.extend(["", "", "", ""])
        # Add extra info for the first member
        member = members[0]
        row.extend([
            member.get(column_mapping.get('gender_identity'), ''),
            member.get(column_mapping.get('sex'), ''),
            member.get(column_mapping.get('residing_ph'), ''),
            member.get(column_mapping.get('gender_preference'), ''),
            member.get(column_mapping.get('country'), ''),
            member.get(column_mapping.get('province'), ''),
            member.get(column_mapping.get('city'), ''),
            member.get(column_mapping.get('state'), ''),
            member.get(column_mapping.get('previous_coach_name'), '')
        ])
        ws.append(row)
        
        # Check if group has 5 or more members and all members have the same location
        should_highlight = False
        if len(members) >= 5:
            # Check if all members have the same location
            first_member = members[0]
            first_residing_ph = str(first_member.get(column_mapping.get('residing_ph'), '0')).strip().lower()
            
            if first_residing_ph in ['1', '1.0', 'true', 'yes', 'ph', 'philippines']:
                # Philippines residents - check if all have same city
                first_city = str(first_member.get(column_mapping.get('city'), '')).strip()
                all_same_location = all(
                    str(member.get(column_mapping.get('city'), '')).strip() == first_city 
                    for member in members[:7]
                )
            else:
                # International residents - check if all have same state and country
                first_state = str(first_member.get(column_mapping.get('state'), '')).strip()
                first_country = str(first_member.get(column_mapping.get('country'), '')).strip()
                all_same_location = all(
                    str(member.get(column_mapping.get('state'), '')).strip() == first_state and
                    str(member.get(column_mapping.get('country'), '')).strip() == first_country
                    for member in members[:7]
                )
            
            if all_same_location:
                regular_group_row_indices.append(ws.max_row)
        
        # Color code user_id and name cells for each member
        for i in range(7):
            if i < len(members):
                member = members[i]
                gender_pref = safe_get_value(member, column_mapping.get('gender_preference', ''), '')
                kaizen_client_type = safe_get_value(member, column_mapping.get('kaizen_client_type', ''), '')
                sex = safe_get_value(member, column_mapping.get('sex', ''), '')
                gender_identity = safe_get_value(member, column_mapping.get('gender_identity', ''), '')
                has_accountability_buddies = safe_get_value(member, column_mapping.get('has_accountability_buddies', ''), '')
                current_goal = safe_get_value(member, column_mapping.get('current_goal', ''), '')
                apply_color_to_cell(ws.cell(row=ws.max_row, column=2 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=True)  # User ID
                apply_color_to_cell(ws.cell(row=ws.max_row, column=3 + i*4), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=False)  # Name
    
    # Apply highlighting to regular groups with 5 or more members and same location
    if regular_group_row_indices:
        # Use a different color for regular groups with 5 or more members and same location (light blue)
        regular_group_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        for row_idx in regular_group_row_indices:
            ws.cell(row=row_idx, column=1).fill = regular_group_fill
    
    # Write excluded users (joiningAsStudent=False)
    if excluded_users:
        for idx, user in enumerate(excluded_users, 1):
            row = [f"Excluded {idx}"]
            
            # Add user data
            location_display = format_location_display(user, column_mapping)
            coach_name = safe_get_value(user, column_mapping.get('previous_coach_name', ''), '')
            age_group = safe_get_value(user, column_mapping.get('age_group', ''), '')
            # Format coach name with age group in parentheses
            coach_with_age = coach_name
            if coach_name and age_group:
                coach_with_age = f"{coach_name} ({age_group})"
            elif coach_name:
                coach_with_age = coach_name
            elif age_group:
                coach_with_age = f"({age_group})"

            row.extend([
                user.get(column_mapping.get('user_id'), ''),
                user.get(column_mapping.get('name'), ''),
                location_display,
                coach_with_age
            ])

            # Add empty cells for remaining slots (to fill up to 7 members)
            for i in range(6):  # 6 more slots (total 7)
                row.extend(["", "", "", ""])
            
            # Add extra info
            row.extend([
                user.get(column_mapping.get('gender_identity'), ''),
                user.get(column_mapping.get('sex'), ''),
                user.get(column_mapping.get('residing_ph'), ''),
                user.get(column_mapping.get('gender_preference'), ''),
                user.get(column_mapping.get('country'), ''),
                user.get(column_mapping.get('province'), ''),
                user.get(column_mapping.get('city'), ''),
                user.get(column_mapping.get('state'), ''),
                user.get(column_mapping.get('previous_coach_name'), '')
            ])
            
            ws.append(row)
            
            # Apply formatting (treat as solo)
            gender_pref = user.get(column_mapping.get('gender_preference'), '')
            kaizen_client_type = user.get(column_mapping.get('kaizen_client_type'), '')
            sex = user.get(column_mapping.get('sex'), '')
            gender_identity = user.get(column_mapping.get('gender_identity'), '')
            has_accountability_buddies = user.get(column_mapping.get('has_accountability_buddies'), '')
            current_goal = user.get(column_mapping.get('current_goal'), '')
            apply_color_to_cell(ws.cell(row=ws.max_row, column=2), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=True)  # User ID
            apply_color_to_cell(ws.cell(row=ws.max_row, column=3), sex, gender_identity, gender_pref, has_accountability_buddies, current_goal, is_user_id=False)  # Name
    
    # Check if filename_or_buffer is a string (file path) or BytesIO buffer
    if isinstance(filename_or_buffer, str):
        wb.save(filename_or_buffer)
    else:
        # It's a BytesIO buffer
        wb.save(filename_or_buffer)

def main():
    """
    MAIN EXECUTION FUNCTION - Complete group assignment pipeline.

    WORKFLOW:
    1. Load merged participant data (Excel/CSV format)
    2. Dynamically detect and map column names
    3. Sort data for consistent processing order
    4. Execute 5-phase grouping algorithm:
       - Accountability buddies (graph-based clustering)
       - Solo participants (user choice)
       - Priority same-gender groups (females first, 5-member target, same location)
       - Regular groups (gender + geography algorithm for remaining)
       - Small group optimization
    5. Generate comprehensive diagnostic reports
    6. Export results to formatted Excel file

    INPUT: Merged participant data file (update INPUT_FILE path)
    OUTPUT: Excel file with organized groups, color coding, and metadata
    """
    # Load input data - supports both Excel and CSV formats
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name='Merged Data')
        print(f"✅ Successfully read input file with {len(df)} records")
    except Exception as e:
        try:
            df = pd.read_csv(INPUT_FILE)
            print(f"✅ Successfully read CSV file with {len(df)} records")
        except Exception as e2:
            print(f"❌ Error reading input file: {e2}")
            return
    
    # Find column mapping
    column_mapping = find_column_mapping(df)
    print(f"\n📋 Column mapping found:")
    for key, value in column_mapping.items():
        if value:
            print(f"  ✅ {key}: {value}")
        else:
            print(f"  ❌ {key}: NOT FOUND")
    
    # Convert DataFrame to list of dictionaries
    # --- SORTING STEP: Sort by province, city, gender_preference, gender_identity, user_id if columns exist ---
    sort_columns = []
    for col_key in ['province', 'city', 'gender_preference', 'gender_identity', 'user_id']:
        col_name = column_mapping.get(col_key)
        if col_name and col_name in df.columns:
            sort_columns.append(col_name)
    if sort_columns:
        df = df.sort_values(by=sort_columns)
        print(f"\n📊 Data sorted by: {sort_columns}")
    data = df.to_dict('records')
    
    print(f"\n🚀 Starting group assignment process...")
    
    # Group participants
    solo_groups, grouped, excluded_users, requested_groups, combined_group_info = group_participants(data, column_mapping)
    
    print(f"\n💾 Saving results to Excel...")
    
    # Save to Excel
    save_to_excel(solo_groups, grouped, OUTPUT_FILE, column_mapping, excluded_users, requested_groups, combined_group_info)
    
    print(f"\n✅ Group assignment completed successfully!")
    print(f"📁 Results saved to: {OUTPUT_FILE}")

# ============================================================================
# EXECUTION ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    # Execute the complete participant grouping pipeline
    # Update INPUT_FILE path in configuration section to point to your merged data file
    main() 